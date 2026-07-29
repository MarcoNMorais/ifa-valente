from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime, time
from functools import wraps
from pathlib import Path
from typing import Any

from flask import (
    Response,
    g,
    jsonify,
    redirect,
    request,
    send_file,
    send_from_directory,
    session,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from werkzeug.security import check_password_hash, generate_password_hash


DEFAULT_UNITS = [
    ("USF Centro", "Centro", "0000001", "centro"),
    ("USF Valilândia", "Valilândia", "0000002", "valilandia"),
    ("USF Casas Populares", "Casas Populares", "0000003", "casaspopulares"),
    ("USF Cidade Nova", "Cidade Nova", "0000004", "cidadenova"),
    ("USF Juazeiro/Petrolina", "Juazeiro/Petrolina", "0000005", "juazeiropetrolina"),
    ("USF Queimada do Curral", "Queimada do Curral", "0000006", "queimada"),
    ("USF Santa Rita de Cássia", "Santa Rita de Cássia", "0000007", "santarita"),
    ("USF Tanquinho", "Tanquinho", "0000008", "tanquinho"),
    ("USF Junco", "Junco", "0000009", "junco"),
    ("USF Dr. Antônio Delfino Mota", "Dr. Antônio", "0000010", "drantonio"),
    ("Secretaria Municipal de Saúde", "Secretaria de Saúde", None, "secretaria"),
    ("CEMES", "CEMES", None, "cemes"),
    ("CIS", "CIS", None, "cis"),
]

PRODUCTION_PROCEDURES = [
    ("Consulta em Urologia", None, "Urologia"),
    ("Ultrassom", None, "Diagnóstico por imagem"),
    ("Consulta em Gastroenterologia", None, "Gastroenterologia"),
    ("OCI em Gastroenterologia", None, "Gastroenterologia"),
    ("OCI em Ortopedia", None, "Ortopedia"),
    ("Consulta em Pediatria", None, "Pediatria"),
    ("Consulta em Ginecologia e Obstetrícia", None, "Ginecologia e Obstetrícia"),
    ("OCI do Colo do Útero", None, "Ginecologia"),
    ("Implanon", None, "Planejamento reprodutivo"),
    ("Consulta em Fonoaudiologia Pediátrica", None, "Fonoaudiologia"),
    ("OCI em Cardiologia", None, "Cardiologia"),
    ("Eletrocardiograma", None, "Cardiologia"),
    ("Consulta em Neurologia", None, "Neurologia"),
    ("Consulta em Reumatologia", None, "Reumatologia"),
]

PRODUCTION_DOCTORS = [
    ("ANDERSON TEIVE E ARGOLLO DULTRA", ["Consulta em Urologia"]),
    ("ANTONIO EDIL MOTA LOPES", ["Ultrassom"]),
    ("AUGUSTO CESAR QUINTELA SOUZA", ["Consulta em Gastroenterologia", "OCI em Gastroenterologia"]),
    ("DILSON LOPES DOS SANTOS", ["OCI em Ortopedia"]),
    ("ENIO HENRIQUE SANTOS RIOS", ["Consulta em Pediatria"]),
    ("FRANCISCO LEON SILVA MASCARENHAS", ["Consulta em Ginecologia e Obstetrícia"]),
    ("GELSON CARNEIRO DA CUNHA", ["Consulta em Ginecologia e Obstetrícia", "OCI do Colo do Útero"]),
    ("IANE MACEDO ARAUJO", ["Implanon"]),
    ("JACKLINE OLIVEIRA FERREIRA BERNARDES", ["Consulta em Fonoaudiologia Pediátrica"]),
    ("LUCAS DE OLIVEIRA ALVES", ["OCI em Gastroenterologia"]),
    ("MICHEL PLATINY MASCARENHAS DE ABREU", ["OCI em Cardiologia"]),
    ("MONIA MARIA CARNEIRO GUIMARAES RAMOS", ["Eletrocardiograma"]),
    ("MURILO CALIXTO DOS SANTOS", ["Consulta em Neurologia"]),
    ("VILSON PATRESE DE JESUS OLIVEIRA", ["Ultrassom"]),
    ("ZENON NUNES DA SILVA FILHO", ["Consulta em Reumatologia"]),
]


SCHEMA = """
CREATE TABLE IF NOT EXISTS units (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE COLLATE NOCASE,
    short_name TEXT,
    cnes TEXT,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    username TEXT NOT NULL UNIQUE COLLATE NOCASE,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('admin','regulacao','unidade','gestor')),
    unit_id INTEGER REFERENCES units(id),
    active INTEGER NOT NULL DEFAULT 1,
    must_change_password INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE UNIQUE INDEX IF NOT EXISTS users_one_unit_profile_idx
ON users(unit_id) WHERE role='unidade' AND active=1;
CREATE TABLE IF NOT EXISTS procedures (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE COLLATE NOCASE,
    sigtap TEXT,
    specialty TEXT,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS doctors (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE COLLATE NOCASE,
    crm TEXT,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS doctor_procedures (
    doctor_id INTEGER NOT NULL REFERENCES doctors(id),
    procedure_id INTEGER NOT NULL REFERENCES procedures(id),
    PRIMARY KEY(doctor_id, procedure_id)
);
CREATE INDEX IF NOT EXISTS doctor_procedures_procedure_idx
ON doctor_procedures(procedure_id, doctor_id);
CREATE TABLE IF NOT EXISTS locations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE COLLATE NOCASE,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS slots (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    unit_id INTEGER NOT NULL REFERENCES units(id),
    procedure_id INTEGER NOT NULL REFERENCES procedures(id),
    service_date TEXT NOT NULL,
    service_time TEXT NOT NULL,
    service_time_max TEXT,
    quantity INTEGER NOT NULL CHECK(quantity > 0),
    doctor_id INTEGER REFERENCES doctors(id),
    provider TEXT,
    location TEXT,
    notes TEXT,
    status TEXT NOT NULL DEFAULT 'active' CHECK(status IN ('active','cancelled','transferred')),
    created_by INTEGER REFERENCES users(id),
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS slots_lookup_idx
ON slots(unit_id, procedure_id, service_date, service_time, status);
CREATE TABLE IF NOT EXISTS bookings (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    slot_id INTEGER NOT NULL REFERENCES slots(id),
    unit_id INTEGER NOT NULL REFERENCES units(id),
    procedure_id INTEGER NOT NULL REFERENCES procedures(id),
    service_date TEXT NOT NULL,
    service_time TEXT NOT NULL,
    operator_id INTEGER REFERENCES users(id),
    source TEXT NOT NULL DEFAULT 'extension' CHECK(source IN ('extension','manual')),
    dedupe_key TEXT NOT NULL UNIQUE,
    status TEXT NOT NULL DEFAULT 'confirmed' CHECK(status IN ('confirmed','cancelled','rescheduled')),
    notes TEXT,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    cancelled_at TEXT,
    cancelled_by INTEGER REFERENCES users(id),
    cancellation_reason TEXT,
    rescheduled_to INTEGER REFERENCES bookings(id)
);
CREATE INDEX IF NOT EXISTS bookings_slot_idx ON bookings(slot_id, status);
CREATE TABLE IF NOT EXISTS pending_records (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    unit_id INTEGER REFERENCES units(id),
    procedure_id INTEGER REFERENCES procedures(id),
    service_date TEXT,
    service_time TEXT,
    operator_id INTEGER REFERENCES users(id),
    dedupe_key TEXT NOT NULL UNIQUE,
    reason TEXT NOT NULL,
    notes TEXT,
    status TEXT NOT NULL DEFAULT 'open' CHECK(status IN ('open','resolved','cancelled')),
    resolution TEXT,
    resolved_by INTEGER REFERENCES users(id),
    resolved_at TEXT,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS extension_devices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    unit_id INTEGER NOT NULL REFERENCES units(id),
    operator_id INTEGER REFERENCES users(id),
    token_hash TEXT NOT NULL UNIQUE,
    active INTEGER NOT NULL DEFAULT 1,
    last_seen_at TEXT,
    created_by INTEGER REFERENCES users(id),
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS audit_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER REFERENCES users(id),
    action TEXT NOT NULL,
    entity TEXT NOT NULL,
    entity_id TEXT,
    before_json TEXT,
    after_json TEXT,
    ip TEXT,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
"""


class CemesError(Exception):
    def __init__(self, message: str, status: int = 400, **extra: Any):
        super().__init__(message)
        self.status = status
        self.extra = extra


def _clean(value: Any, limit: int = 200) -> str | None:
    text = str(value if value is not None else "").strip()
    return text[:limit] or None


def _active(value: Any, fallback: bool = True) -> bool:
    if value in (None, ""):
        return bool(fallback)
    return value in (True, 1, "1", "true", "True")


def _date(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value or "")[:10]
    try:
        datetime.strptime(text, "%Y-%m-%d")
        return text
    except ValueError:
        return None


def _time(value: Any) -> str | None:
    if isinstance(value, (datetime, time)):
        return value.strftime("%H:%M")
    match = re.match(r"^(\d{2}):(\d{2})", str(value or ""))
    if not match or int(match.group(1)) > 23 or int(match.group(2)) > 59:
        return None
    return f"{match.group(1)}:{match.group(2)}"


def _norm(value: Any) -> str:
    import unicodedata

    return "".join(
        char for char in unicodedata.normalize("NFD", str(value or "").lower())
        if unicodedata.category(char) != "Mn"
    ).strip()


def _row(row: sqlite3.Row | None) -> dict[str, Any] | None:
    return dict(row) if row is not None else None


def _rows(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    return [dict(row) for row in rows]


class CemesStore:
    def __init__(self, app_root: str):
        if os.environ.get("CEMES_DATA_DIR"):
            self.data_dir = Path(os.environ["CEMES_DATA_DIR"])
        elif os.environ.get("IFA_DATA_DIR"):
            self.data_dir = Path(os.environ["IFA_DATA_DIR"]) / "cemes"
        else:
            self.data_dir = Path(app_root) / "data" / "cemes"
        self.backup_dir = Path(os.environ.get("CEMES_BACKUP_DIR", self.data_dir / "backups"))
        self.db_path = self.data_dir / "cmvr.db"
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.initialize()

    def connect(self) -> sqlite3.Connection:
        db = sqlite3.connect(self.db_path, timeout=10)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys=ON")
        db.execute("PRAGMA journal_mode=WAL")
        db.execute("PRAGMA busy_timeout=5000")
        return db

    @contextmanager
    def connection(self):
        db = self.connect()
        try:
            with db:
                yield db
        finally:
            db.close()

    def initialize(self) -> None:
        with self.connection() as db:
            db.executescript(SCHEMA)
            base_marker = db.execute(
                "SELECT 1 FROM settings WHERE key='seed:v1:base-catalog'"
            ).fetchone()
            if not base_marker:
                if db.execute("SELECT COUNT(*) FROM units").fetchone()[0] == 0:
                    for name, short_name, cnes, _username in DEFAULT_UNITS:
                        db.execute(
                            "INSERT INTO units(name,short_name,cnes) VALUES(?,?,?)",
                            (name, short_name, cnes),
                        )
                if db.execute("SELECT COUNT(*) FROM procedures").fetchone()[0] == 0:
                    for name, sigtap, specialty in PRODUCTION_PROCEDURES:
                        db.execute(
                            "INSERT INTO procedures(name,sigtap,specialty) VALUES(?,?,?)",
                            (name, sigtap, specialty),
                        )
                if db.execute("SELECT COUNT(*) FROM locations").fetchone()[0] == 0:
                    db.execute("INSERT INTO locations(name) VALUES('CEMES')")
                db.execute(
                    "INSERT INTO settings(key,value) VALUES('seed:v1:base-catalog',?)",
                    (json.dumps({"units": 13, "procedures": 14, "locations": 1}),),
                )

            marker = db.execute(
                "SELECT 1 FROM settings WHERE key='catalog:v1:medical-roster'"
            ).fetchone()
            if not marker:
                for doctor_name, procedure_names in PRODUCTION_DOCTORS:
                    db.execute("INSERT OR IGNORE INTO doctors(name) VALUES(?)", (doctor_name,))
                    doctor_id = db.execute(
                        "SELECT id FROM doctors WHERE name=? COLLATE NOCASE", (doctor_name,)
                    ).fetchone()["id"]
                    for procedure_name in procedure_names:
                        procedure_id = db.execute(
                            "SELECT id FROM procedures WHERE name=? COLLATE NOCASE",
                            (procedure_name,),
                        ).fetchone()["id"]
                        db.execute(
                            "INSERT OR IGNORE INTO doctor_procedures(doctor_id,procedure_id) VALUES(?,?)",
                            (doctor_id, procedure_id),
                        )
                db.execute(
                    "INSERT INTO settings(key,value) VALUES('catalog:v1:medical-roster',?)",
                    (json.dumps({"doctors": 15, "procedures": 14}),),
                )

            if db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
                initial_users = [
                    (
                        "Administrador",
                        "admin",
                        os.environ.get("CMVR_ADMIN_PASSWORD", "Admin@123"),
                        "admin",
                        None,
                    ),
                    (
                        "Regulação Central",
                        "regulacao",
                        os.environ.get("CMVR_REGULACAO_PASSWORD", "Regulacao@123"),
                        "regulacao",
                        None,
                    ),
                    (
                        "Gestor Municipal",
                        "gestor",
                        os.environ.get("CMVR_GESTOR_PASSWORD", "Gestor@123"),
                        "gestor",
                        None,
                    ),
                ]
                unit_password = os.environ.get("CMVR_UNIDADE_PASSWORD", "Unidade@123")
                for unit_name, short_name, _cnes, username in DEFAULT_UNITS:
                    unit_id = db.execute(
                        "SELECT id FROM units WHERE name=? COLLATE NOCASE", (unit_name,)
                    ).fetchone()["id"]
                    initial_users.append(
                        (f"Perfil {short_name}", username, unit_password, "unidade", unit_id)
                    )
                for name, username, password, role, unit_id in initial_users:
                    db.execute(
                        """
                        INSERT INTO users(name,username,password_hash,role,unit_id)
                        VALUES(?,?,?,?,?)
                        """,
                        (name, username, generate_password_hash(password), role, unit_id),
                    )

    def audit(
        self,
        db: sqlite3.Connection,
        action: str,
        entity: str,
        entity_id: Any = None,
        user_id: int | None = None,
        before: Any = None,
        after: Any = None,
        ip: str | None = None,
    ) -> None:
        db.execute(
            """
            INSERT INTO audit_logs(user_id,action,entity,entity_id,before_json,after_json,ip)
            VALUES(?,?,?,?,?,?,?)
            """,
            (
                user_id,
                action,
                entity,
                None if entity_id is None else str(entity_id),
                None if before is None else json.dumps(before, ensure_ascii=False, default=str),
                None if after is None else json.dumps(after, ensure_ascii=False, default=str),
                ip,
            ),
        )

    def list_backups(self) -> list[dict[str, Any]]:
        items = []
        for path in self.backup_dir.glob("*.db"):
            if re.match(r"^(auto|manual)-[\w.-]+\.db$", path.name):
                stat = path.stat()
                items.append(
                    {
                        "name": path.name,
                        "size": stat.st_size,
                        "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    }
                )
        return sorted(items, key=lambda item: item["created_at"], reverse=True)

    def create_backup(self, prefix: str = "manual") -> dict[str, Any]:
        stamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S-%f")
        target = self.backup_dir / f"{prefix}-{stamp}.db"
        source_db = sqlite3.connect(self.db_path)
        target_db = sqlite3.connect(target)
        try:
            source_db.backup(target_db)
        finally:
            target_db.close()
            source_db.close()
        backups = self.list_backups()
        for old in backups[10:]:
            (self.backup_dir / old["name"]).unlink(missing_ok=True)
        return {"name": target.name, "size": target.stat().st_size}

    def auto_backup(self) -> None:
        today = datetime.now().strftime("%Y-%m-%d")
        if not any(item["name"].startswith(f"auto-{today}") for item in self.list_backups()):
            self.create_backup("auto")

    def backup_path(self, name: str) -> Path | None:
        if not re.match(r"^(auto|manual)-[\w.-]+\.db$", str(name)):
            return None
        path = self.backup_dir / Path(name).name
        return path if path.is_file() else None

    def restore_backup(self, name: str) -> dict[str, Any]:
        source = self.backup_path(name)
        if not source:
            raise CemesError("Backup não encontrado.", 404)
        safety = self.create_backup("manual")
        source_db = sqlite3.connect(source)
        target_db = sqlite3.connect(self.db_path)
        try:
            source_db.backup(target_db)
        finally:
            target_db.close()
            source_db.close()
        self.initialize()
        return {"restored": name, "safety_backup": safety["name"]}


def register_cemes_routes(app):
    store = CemesStore(app.root_path)
    cemes_dir = Path(app.root_path) / "Cemes"
    login_attempts: dict[str, tuple[int, float]] = {}

    def api_guard(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            try:
                return view(*args, **kwargs)
            except CemesError as error:
                return jsonify(error=str(error), **error.extra), error.status
            except sqlite3.IntegrityError as error:
                app.logger.warning("CMVR integrity error: %s", error)
                return jsonify(error="Já existe um registro com essas informações."), 409
            except Exception:
                app.logger.exception("Falha no módulo CEMES")
                return jsonify(error="Não foi possível concluir a operação."), 500

        return wrapped

    def permissions(user: dict[str, Any]) -> dict[str, Any]:
        is_cemes = user["role"] == "unidade" and _norm(user.get("unit_name")) == "cemes"
        return {
            "view_all_units": user["role"] != "unidade" or is_cemes,
            "manual_booking_scope": (
                "all"
                if user["role"] in ("admin", "regulacao")
                else "own_and_secretaria"
                if is_cemes
                else "own"
                if user["role"] == "unidade"
                else "none"
            ),
        }

    def current_user() -> dict[str, Any] | None:
        user_id = session.get("cmvr_user_id")
        if not user_id:
            return None
        with store.connection() as db:
            row = db.execute(
                """
                SELECT usr.id,usr.name,usr.username,usr.role,usr.unit_id,usr.active,
                       usr.must_change_password,u.name unit_name
                FROM users usr LEFT JOIN units u ON u.id=usr.unit_id
                WHERE usr.id=? AND usr.active=1
                """,
                (user_id,),
            ).fetchone()
        if not row:
            for key in ("cmvr_user_id", "cmvr_csrf"):
                session.pop(key, None)
            return None
        user = dict(row)
        user["must_change_password"] = bool(user["must_change_password"])
        user["permissions"] = permissions(user)
        return user

    def require_auth(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = current_user()
            if not user:
                raise CemesError("Faça login para continuar.", 401)
            g.cmvr_user = user
            return view(*args, **kwargs)

        return wrapped

    def require_roles(*roles):
        def decorator(view):
            @wraps(view)
            def wrapped(*args, **kwargs):
                if g.cmvr_user["role"] not in roles:
                    raise CemesError("Você não possui permissão para esta operação.", 403)
                return view(*args, **kwargs)

            return wrapped

        return decorator

    def require_csrf(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            supplied = request.headers.get("X-CSRF-Token", "")
            expected = session.get("cmvr_csrf", "")
            if not expected or not secrets.compare_digest(supplied, expected):
                raise CemesError("Sessão de segurança inválida. Atualize a página.", 403)
            return view(*args, **kwargs)

        return wrapped

    def body() -> dict[str, Any]:
        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            raise CemesError("Conteúdo inválido.")
        return data

    def scoped_unit_id(user: dict[str, Any]) -> int | None:
        if user["permissions"]["view_all_units"]:
            value = request.args.get("unit_id")
            return int(value) if value and value.isdigit() else None
        return int(user["unit_id"])

    def slot_by_id(db: sqlite3.Connection, slot_id: int) -> dict[str, Any] | None:
        return _row(
            db.execute(
                """
                SELECT s.*,u.name unit_name,p.name procedure_name,p.sigtap,
                       COALESCE(d.name,s.provider) doctor_name,
                       (SELECT COUNT(*) FROM bookings b
                        WHERE b.slot_id=s.id AND b.status='confirmed') used,
                       s.quantity-(SELECT COUNT(*) FROM bookings b
                        WHERE b.slot_id=s.id AND b.status='confirmed') remaining
                FROM slots s JOIN units u ON u.id=s.unit_id
                JOIN procedures p ON p.id=s.procedure_id
                LEFT JOIN doctors d ON d.id=s.doctor_id
                WHERE s.id=?
                """,
                (slot_id,),
            ).fetchone()
        )

    def booking_by_id(db: sqlite3.Connection, booking_id: int) -> dict[str, Any] | None:
        return _row(
            db.execute(
                """
                SELECT b.*,u.name unit_name,p.name procedure_name,usr.name operator_name
                FROM bookings b JOIN units u ON u.id=b.unit_id
                JOIN procedures p ON p.id=b.procedure_id
                LEFT JOIN users usr ON usr.id=b.operator_id WHERE b.id=?
                """,
                (booking_id,),
            ).fetchone()
        )

    def list_slots(
        db: sqlite3.Connection,
        unit_id: int | None = None,
        procedure_id: Any = None,
        date_from: Any = None,
        date_to: Any = None,
        available_only: bool = False,
    ) -> list[dict[str, Any]]:
        where = ["s.status='active'"]
        params: list[Any] = []
        if unit_id:
            where.append("s.unit_id=?")
            params.append(unit_id)
        if procedure_id:
            where.append("s.procedure_id=?")
            params.append(int(procedure_id))
        if date_from:
            where.append("s.service_date>=?")
            params.append(str(date_from))
        if date_to:
            where.append("s.service_date<=?")
            params.append(str(date_to))
        if available_only:
            where.append(
                "(SELECT COUNT(*) FROM bookings b WHERE b.slot_id=s.id AND b.status='confirmed') < s.quantity"
            )
        return _rows(
            db.execute(
                f"""
                SELECT s.*,u.name unit_name,p.name procedure_name,p.sigtap,
                       COALESCE(d.name,s.provider) doctor_name,
                       (SELECT COUNT(*) FROM bookings b
                        WHERE b.slot_id=s.id AND b.status='confirmed') used,
                       s.quantity-(SELECT COUNT(*) FROM bookings b
                        WHERE b.slot_id=s.id AND b.status='confirmed') remaining
                FROM slots s JOIN units u ON u.id=s.unit_id
                JOIN procedures p ON p.id=s.procedure_id
                LEFT JOIN doctors d ON d.id=s.doctor_id
                WHERE {" AND ".join(where)}
                ORDER BY s.service_date,s.service_time,p.name,u.name
                """,
                params,
            ).fetchall()
        )

    def validate_doctor(
        db: sqlite3.Connection, doctor_id: int, procedure_id: int
    ) -> dict[str, Any]:
        doctor = _row(
            db.execute(
                """
                SELECT d.id,d.name FROM doctors d
                JOIN doctor_procedures dp ON dp.doctor_id=d.id
                WHERE d.id=? AND dp.procedure_id=? AND d.active=1
                """,
                (doctor_id, procedure_id),
            ).fetchone()
        )
        if not doctor:
            raise CemesError("O médico selecionado não está vinculado a este procedimento.")
        return doctor

    def insert_slot(
        db: sqlite3.Connection,
        data: dict[str, Any],
        user_id: int,
        ip: str,
        do_audit: bool = True,
    ) -> dict[str, Any]:
        service_date = _date(data.get("service_date"))
        service_time = _time(data.get("service_time"))
        service_time_max = _time(data.get("service_time_max") or data.get("service_time"))
        try:
            quantity = int(data.get("quantity"))
            unit_id = int(data.get("unit_id"))
            procedure_id = int(data.get("procedure_id"))
        except (TypeError, ValueError):
            raise CemesError("Data, horário ou quantidade inválidos.")
        if (
            not service_date
            or not service_time
            or not service_time_max
            or service_time_max < service_time
            or quantity < 1
            or quantity > 5000
        ):
            raise CemesError("Data, horário ou quantidade inválidos.")
        if not db.execute("SELECT 1 FROM units WHERE id=? AND active=1", (unit_id,)).fetchone():
            raise CemesError("A unidade informada não existe ou está inativa.")
        if not db.execute(
            "SELECT 1 FROM procedures WHERE id=? AND active=1", (procedure_id,)
        ).fetchone():
            raise CemesError("O procedimento informado não existe ou está inativo.")
        doctor_id = int(data.get("doctor_id") or 0) or None
        provider = _clean(data.get("provider"))
        if doctor_id:
            doctor = validate_doctor(db, doctor_id, procedure_id)
            provider = doctor["name"]
        cursor = db.execute(
            """
            INSERT INTO slots(
                unit_id,procedure_id,service_date,service_time,service_time_max,
                quantity,doctor_id,provider,location,notes,created_by
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                unit_id,
                procedure_id,
                service_date,
                service_time,
                service_time_max,
                quantity,
                doctor_id,
                provider,
                _clean(data.get("location")) or "CEMES",
                _clean(data.get("notes"), 500),
                user_id,
            ),
        )
        created = slot_by_id(db, cursor.lastrowid)
        if do_audit:
            store.audit(
                db,
                "CREATE",
                "slot",
                cursor.lastrowid,
                user_id=user_id,
                after=created,
                ip=ip,
            )
        return created

    def confirm_booking(
        db: sqlite3.Connection,
        data: dict[str, Any],
        unit_id: int,
        user_id: int | None,
        ip: str,
        source: str,
    ) -> tuple[dict[str, Any], int]:
        service_date = _date(data.get("service_date"))
        service_time = _time(data.get("service_time"))
        try:
            procedure_id = int(data.get("procedure_id"))
        except (TypeError, ValueError):
            procedure_id = 0
        dedupe_key = str(data.get("dedupe_key") or "")[:180]
        if source == "manual" and len(dedupe_key) < 8:
            dedupe_key = f"manual-{secrets.token_urlsafe(18)}"
        if not service_date or not service_time or not procedure_id or len(dedupe_key) < 8:
            raise CemesError("Procedimento, data, horário e identificador são obrigatórios.")
        existing = _row(
            db.execute("SELECT * FROM bookings WHERE dedupe_key=?", (dedupe_key,)).fetchone()
        )
        if existing:
            return (
                {
                    "status": "duplicate",
                    "booking": existing,
                    "message": "Este agendamento já havia sido registrado.",
                },
                200,
            )
        pending_existing = _row(
            db.execute(
                "SELECT * FROM pending_records WHERE dedupe_key=?", (dedupe_key,)
            ).fetchone()
        )
        if pending_existing:
            return (
                {
                    "status": "duplicate_pending",
                    "pending": pending_existing,
                    "message": "Esta ocorrência já está em pendência.",
                },
                200,
            )
        slot = _row(
            db.execute(
                """
                SELECT s.*,(SELECT COUNT(*) FROM bookings b
                WHERE b.slot_id=s.id AND b.status='confirmed') used
                FROM slots s
                WHERE s.unit_id=? AND s.procedure_id=? AND s.service_date=?
                  AND s.service_time<=?
                  AND COALESCE(s.service_time_max,s.service_time)>=?
                  AND s.status='active'
                  AND (SELECT COUNT(*) FROM bookings b
                       WHERE b.slot_id=s.id AND b.status='confirmed') < s.quantity
                ORDER BY s.service_time DESC,s.id LIMIT 1
                """,
                (unit_id, procedure_id, service_date, service_time, service_time),
            ).fetchone()
        )
        if not slot:
            matching = db.execute(
                """
                SELECT 1 FROM slots
                WHERE unit_id=? AND procedure_id=? AND service_date=?
                  AND service_time<=?
                  AND COALESCE(service_time_max,service_time)>=?
                  AND status='active' LIMIT 1
                """,
                (unit_id, procedure_id, service_date, service_time, service_time),
            ).fetchone()
            reason = (
                "Horário sem saldo disponível."
                if matching
                else "Vaga não cadastrada para esta unidade, procedimento, data e horário."
            )
            cursor = db.execute(
                """
                INSERT INTO pending_records(
                    unit_id,procedure_id,service_date,service_time,operator_id,
                    dedupe_key,reason,notes
                ) VALUES(?,?,?,?,?,?,?,?)
                """,
                (
                    unit_id,
                    procedure_id,
                    service_date,
                    service_time,
                    user_id,
                    dedupe_key,
                    reason,
                    _clean(data.get("notes"), 400),
                ),
            )
            pending = _row(
                db.execute(
                    "SELECT * FROM pending_records WHERE id=?", (cursor.lastrowid,)
                ).fetchone()
            )
            store.audit(
                db,
                "CREATE_PENDING",
                "pending",
                cursor.lastrowid,
                user_id=user_id,
                after=pending,
                ip=ip,
            )
            return {"status": "pending", "pending": pending, "message": reason}, 202
        cursor = db.execute(
            """
            INSERT INTO bookings(
                slot_id,unit_id,procedure_id,service_date,service_time,
                operator_id,source,dedupe_key,notes
            ) VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (
                slot["id"],
                unit_id,
                procedure_id,
                service_date,
                service_time,
                user_id,
                "manual" if source == "manual" else "extension",
                dedupe_key,
                _clean(data.get("notes"), 400),
            ),
        )
        booking = booking_by_id(db, cursor.lastrowid)
        store.audit(
            db,
            "CONFIRM",
            "booking",
            cursor.lastrowid,
            user_id=user_id,
            after=booking,
            ip=ip,
        )
        return (
            {
                "status": "confirmed",
                "booking": booking,
                "message": "Vaga registrada com sucesso.",
            },
            201 if source == "manual" else 200,
        )

    @app.after_request
    def cemes_extension_cors(response: Response) -> Response:
        if request.path.startswith("/Cemes/api/extension"):
            response.headers["Access-Control-Allow-Origin"] = "*"
            response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        return response

    @app.route("/Cemes/api/health")
    def cemes_health():
        return jsonify(
            status="ok",
            service="CMVR",
            storage="persistent" if os.environ.get("IFA_DATA_DIR") else "local",
            time=datetime.now().isoformat(),
        )

    @app.route("/Cemes/api/auth/login", methods=["POST"])
    @api_guard
    def cemes_login():
        import time as time_module

        data = body()
        username = _clean(data.get("username"), 80) or ""
        key = f"{request.remote_addr}:{username.lower()}"
        count, last = login_attempts.get(key, (0, 0))
        if count >= 5 and time_module.time() - last < 900:
            raise CemesError("Muitas tentativas. Aguarde 15 minutos antes de tentar novamente.", 429)
        with store.connection() as db:
            user = _row(
                db.execute(
                    "SELECT * FROM users WHERE username=? COLLATE NOCASE AND active=1",
                    (username,),
                ).fetchone()
            )
            if not user or not check_password_hash(
                user["password_hash"], str(data.get("password") or "")
            ):
                login_attempts[key] = (count + 1, time_module.time())
                store.audit(
                    db,
                    "LOGIN_FAILED",
                    "session",
                    after={"username": username},
                    ip=request.remote_addr,
                )
                raise CemesError("Usuário ou senha inválidos.", 401)
            login_attempts.pop(key, None)
            session["cmvr_user_id"] = user["id"]
            session["cmvr_csrf"] = secrets.token_urlsafe(32)
            session.permanent = True
            store.audit(
                db,
                "LOGIN",
                "session",
                user["id"],
                user_id=user["id"],
                ip=request.remote_addr,
            )
        return jsonify(ok=True)

    @app.route("/Cemes/api/auth/logout", methods=["POST"])
    @api_guard
    @require_auth
    @require_csrf
    def cemes_logout():
        with store.connection() as db:
            store.audit(
                db,
                "LOGOUT",
                "session",
                g.cmvr_user["id"],
                user_id=g.cmvr_user["id"],
                ip=request.remote_addr,
            )
        session.pop("cmvr_user_id", None)
        session.pop("cmvr_csrf", None)
        return jsonify(ok=True)

    @app.route("/Cemes/api/session")
    @api_guard
    @require_auth
    def cemes_session():
        return jsonify(user=g.cmvr_user, csrf_token=session["cmvr_csrf"])

    @app.route("/Cemes/api/dashboard")
    @api_guard
    @require_auth
    def cemes_dashboard():
        unit_id = scoped_unit_id(g.cmvr_user)
        clause = " AND s.unit_id=?" if unit_id else ""
        params = [unit_id] if unit_id else []
        with store.connection() as db:
            totals = db.execute(
                f"""
                SELECT COALESCE(SUM(s.quantity),0) total,
                COALESCE(SUM((SELECT COUNT(*) FROM bookings b
                WHERE b.slot_id=s.id AND b.status='confirmed')),0) used
                FROM slots s WHERE s.status='active'{clause}
                """,
                params,
            ).fetchone()
            pending = db.execute(
                "SELECT COUNT(*) FROM pending_records WHERE status='open'"
                + (" AND unit_id=?" if unit_id else ""),
                params,
            ).fetchone()[0]
            expired = db.execute(
                f"""
                SELECT COALESCE(SUM(MAX(0,s.quantity-(SELECT COUNT(*) FROM bookings b
                WHERE b.slot_id=s.id AND b.status='confirmed'))),0)
                FROM slots s WHERE s.status='active'
                AND datetime(s.service_date||' '||s.service_time)<datetime('now','localtime')
                {clause}
                """,
                params,
            ).fetchone()[0]
            today = db.execute(
                "SELECT COUNT(*) FROM bookings b WHERE date(b.created_at,'localtime')=date('now','localtime') "
                "AND b.status='confirmed'" + (" AND b.unit_id=?" if unit_id else ""),
                params,
            ).fetchone()[0]
            by_procedure = _rows(
                db.execute(
                    f"""
                    SELECT p.name,SUM(s.quantity) total,
                    SUM((SELECT COUNT(*) FROM bookings b
                    WHERE b.slot_id=s.id AND b.status='confirmed')) used
                    FROM slots s JOIN procedures p ON p.id=s.procedure_id
                    WHERE s.status='active'{clause}
                    GROUP BY p.id ORDER BY used DESC,p.name LIMIT 8
                    """,
                    params,
                ).fetchall()
            )
        return jsonify(
            total=totals["total"],
            used=totals["used"],
            remaining=totals["total"] - totals["used"],
            pending=pending,
            expired=expired,
            today=today,
            byProcedure=by_procedure,
        )

    @app.route("/Cemes/api/units", methods=["GET", "POST"])
    @api_guard
    @require_auth
    def cemes_units():
        if request.method == "GET":
            include_inactive = (
                request.args.get("include_inactive") == "1"
                and g.cmvr_user["permissions"]["view_all_units"]
                and g.cmvr_user["role"] in ("admin", "regulacao")
            )
            with store.connection() as db:
                if g.cmvr_user["permissions"]["view_all_units"]:
                    query = "SELECT * FROM units"
                    if not include_inactive:
                        query += " WHERE active=1"
                    query += " ORDER BY active DESC,name"
                    result = _rows(db.execute(query).fetchall())
                else:
                    result = _rows(
                        db.execute(
                            "SELECT * FROM units WHERE id=?", (g.cmvr_user["unit_id"],)
                        ).fetchall()
                    )
            return jsonify(result)
        if g.cmvr_user["role"] != "admin":
            raise CemesError("Você não possui permissão para esta operação.", 403)
        require_csrf(lambda: None)()
        data = body()
        name = _clean(data.get("name"), 160)
        if not name:
            raise CemesError("Informe o nome da unidade.")
        store.auto_backup()
        with store.connection() as db:
            cursor = db.execute(
                "INSERT INTO units(name,short_name,cnes) VALUES(?,?,?)",
                (name, _clean(data.get("short_name"), 80), _clean(data.get("cnes"), 20)),
            )
            unit = _row(db.execute("SELECT * FROM units WHERE id=?", (cursor.lastrowid,)).fetchone())
            store.audit(
                db,
                "CREATE",
                "unit",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after=unit,
                ip=request.remote_addr,
            )
        return jsonify(unit), 201

    @app.route("/Cemes/api/units/<int:unit_id>", methods=["PUT"])
    @api_guard
    @require_auth
    @require_roles("admin")
    @require_csrf
    def cemes_unit_update(unit_id: int):
        data = body()
        store.auto_backup()
        with store.connection() as db:
            before = _row(db.execute("SELECT * FROM units WHERE id=?", (unit_id,)).fetchone())
            if not before:
                raise CemesError("Unidade não encontrada.", 404)
            name = _clean(data.get("name"), 160)
            if not name:
                raise CemesError("Informe o nome da unidade.")
            active = _active(data.get("active"), bool(before["active"]))
            db.execute(
                "UPDATE units SET name=?,short_name=?,cnes=?,active=? WHERE id=?",
                (
                    name,
                    _clean(data.get("short_name"), 80),
                    _clean(data.get("cnes"), 20),
                    1 if active else 0,
                    unit_id,
                ),
            )
            if active:
                profile = db.execute(
                    "SELECT id FROM users WHERE unit_id=? AND role='unidade' ORDER BY id LIMIT 1",
                    (unit_id,),
                ).fetchone()
                db.execute(
                    "UPDATE users SET active=0 WHERE unit_id=? AND role='unidade'", (unit_id,)
                )
                if profile:
                    db.execute("UPDATE users SET active=1 WHERE id=?", (profile["id"],))
                db.execute("UPDATE extension_devices SET active=1 WHERE unit_id=?", (unit_id,))
            else:
                db.execute(
                    "UPDATE users SET active=0 WHERE unit_id=? AND role='unidade'", (unit_id,)
                )
                db.execute("UPDATE extension_devices SET active=0 WHERE unit_id=?", (unit_id,))
            after = _row(db.execute("SELECT * FROM units WHERE id=?", (unit_id,)).fetchone())
            store.audit(
                db,
                "UPDATE",
                "unit",
                unit_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/procedures", methods=["GET", "POST"])
    @api_guard
    @require_auth
    def cemes_procedures():
        if request.method == "GET":
            include = (
                request.args.get("include_inactive") == "1"
                and g.cmvr_user["role"] in ("admin", "regulacao")
            )
            with store.connection() as db:
                rows = _rows(
                    db.execute(
                        "SELECT * FROM procedures "
                        + ("" if include else "WHERE active=1 ")
                        + "ORDER BY active DESC,name"
                    ).fetchall()
                )
            return jsonify(rows)
        if g.cmvr_user["role"] not in ("admin", "regulacao"):
            raise CemesError("Você não possui permissão para esta operação.", 403)
        require_csrf(lambda: None)()
        data = body()
        name = _clean(data.get("name"), 180)
        if not name:
            raise CemesError("Informe o procedimento.")
        store.auto_backup()
        with store.connection() as db:
            cursor = db.execute(
                "INSERT INTO procedures(name,sigtap,specialty) VALUES(?,?,?)",
                (name, _clean(data.get("sigtap"), 20), _clean(data.get("specialty"), 100)),
            )
            procedure = _row(
                db.execute("SELECT * FROM procedures WHERE id=?", (cursor.lastrowid,)).fetchone()
            )
            store.audit(
                db,
                "CREATE",
                "procedure",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after=procedure,
                ip=request.remote_addr,
            )
        return jsonify(procedure), 201

    @app.route("/Cemes/api/procedures/<int:procedure_id>", methods=["PUT"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_procedure_update(procedure_id: int):
        data = body()
        store.auto_backup()
        with store.connection() as db:
            before = _row(
                db.execute("SELECT * FROM procedures WHERE id=?", (procedure_id,)).fetchone()
            )
            if not before:
                raise CemesError("Procedimento não encontrado.", 404)
            name = _clean(data.get("name"), 180)
            if not name:
                raise CemesError("Informe o procedimento.")
            db.execute(
                "UPDATE procedures SET name=?,sigtap=?,specialty=?,active=? WHERE id=?",
                (
                    name,
                    _clean(data.get("sigtap"), 20),
                    _clean(data.get("specialty"), 100),
                    1 if _active(data.get("active"), bool(before["active"])) else 0,
                    procedure_id,
                ),
            )
            after = _row(
                db.execute("SELECT * FROM procedures WHERE id=?", (procedure_id,)).fetchone()
            )
            store.audit(
                db,
                "UPDATE",
                "procedure",
                procedure_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/doctors", methods=["GET", "POST"])
    @api_guard
    @require_auth
    def cemes_doctors():
        if request.method == "GET":
            include = (
                request.args.get("include_inactive") == "1"
                and g.cmvr_user["role"] in ("admin", "regulacao")
            )
            procedure_id = request.args.get("procedure_id")
            params: list[Any] = []
            where = "1=1" if include else "d.active=1"
            if procedure_id:
                where += (
                    " AND EXISTS(SELECT 1 FROM doctor_procedures f "
                    "WHERE f.doctor_id=d.id AND f.procedure_id=?)"
                )
                params.append(int(procedure_id))
            with store.connection() as db:
                doctors = _rows(
                    db.execute(
                        f"SELECT d.* FROM doctors d WHERE {where} ORDER BY d.active DESC,d.name",
                        params,
                    ).fetchall()
                )
                links = _rows(
                    db.execute(
                        """
                        SELECT dp.doctor_id,dp.procedure_id,p.name procedure_name
                        FROM doctor_procedures dp JOIN procedures p ON p.id=dp.procedure_id
                        ORDER BY p.name
                        """
                    ).fetchall()
                )
            for doctor in doctors:
                own = [link for link in links if link["doctor_id"] == doctor["id"]]
                doctor["procedure_ids"] = [link["procedure_id"] for link in own]
                doctor["procedure_names"] = [link["procedure_name"] for link in own]
            return jsonify(doctors)
        if g.cmvr_user["role"] not in ("admin", "regulacao"):
            raise CemesError("Você não possui permissão para esta operação.", 403)
        require_csrf(lambda: None)()
        data = body()
        name = _clean(data.get("name"), 180)
        ids = sorted({int(value) for value in data.get("procedure_ids", []) if str(value).isdigit()})
        if not name:
            raise CemesError("Informe o nome do médico.")
        if not ids:
            raise CemesError("Selecione ao menos um procedimento para o médico.")
        store.auto_backup()
        with store.connection() as db:
            placeholders = ",".join("?" for _ in ids)
            valid = db.execute(
                f"SELECT COUNT(*) FROM procedures WHERE active=1 AND id IN ({placeholders})",
                ids,
            ).fetchone()[0]
            if valid != len(ids):
                raise CemesError("Um dos procedimentos selecionados não está disponível.")
            cursor = db.execute(
                "INSERT INTO doctors(name,crm) VALUES(?,?)",
                (name, _clean(data.get("crm"), 40)),
            )
            for procedure_id in ids:
                db.execute(
                    "INSERT INTO doctor_procedures(doctor_id,procedure_id) VALUES(?,?)",
                    (cursor.lastrowid, procedure_id),
                )
            doctor = _row(
                db.execute("SELECT * FROM doctors WHERE id=?", (cursor.lastrowid,)).fetchone()
            )
            doctor["procedure_ids"] = ids
            store.audit(
                db,
                "CREATE",
                "doctor",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after=doctor,
                ip=request.remote_addr,
            )
        return jsonify(doctor), 201

    @app.route("/Cemes/api/doctors/<int:doctor_id>", methods=["PUT"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_doctor_update(doctor_id: int):
        data = body()
        name = _clean(data.get("name"), 180)
        ids = sorted({int(value) for value in data.get("procedure_ids", []) if str(value).isdigit()})
        if not name:
            raise CemesError("Informe o nome do médico.")
        if not ids:
            raise CemesError("Selecione ao menos um procedimento para o médico.")
        store.auto_backup()
        with store.connection() as db:
            before = _row(db.execute("SELECT * FROM doctors WHERE id=?", (doctor_id,)).fetchone())
            if not before:
                raise CemesError("Médico não encontrado.", 404)
            placeholders = ",".join("?" for _ in ids)
            valid = db.execute(
                f"SELECT COUNT(*) FROM procedures WHERE id IN ({placeholders})", ids
            ).fetchone()[0]
            if valid != len(ids):
                raise CemesError("Um dos procedimentos selecionados não existe.")
            db.execute(
                "UPDATE doctors SET name=?,crm=?,active=? WHERE id=?",
                (
                    name,
                    _clean(data.get("crm"), 40),
                    1 if _active(data.get("active"), bool(before["active"])) else 0,
                    doctor_id,
                ),
            )
            db.execute("DELETE FROM doctor_procedures WHERE doctor_id=?", (doctor_id,))
            for procedure_id in ids:
                db.execute(
                    "INSERT INTO doctor_procedures(doctor_id,procedure_id) VALUES(?,?)",
                    (doctor_id, procedure_id),
                )
            after = _row(db.execute("SELECT * FROM doctors WHERE id=?", (doctor_id,)).fetchone())
            after["procedure_ids"] = ids
            store.audit(
                db,
                "UPDATE",
                "doctor",
                doctor_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/locations", methods=["GET", "POST"])
    @api_guard
    @require_auth
    def cemes_locations():
        if request.method == "GET":
            with store.connection() as db:
                return jsonify(
                    _rows(
                        db.execute(
                            "SELECT * FROM locations WHERE active=1 ORDER BY name"
                        ).fetchall()
                    )
                )
        if g.cmvr_user["role"] not in ("admin", "regulacao"):
            raise CemesError("Você não possui permissão para esta operação.", 403)
        require_csrf(lambda: None)()
        name = _clean(body().get("name"), 180)
        if not name:
            raise CemesError("Informe o local de atendimento.")
        store.auto_backup()
        with store.connection() as db:
            cursor = db.execute("INSERT INTO locations(name) VALUES(?)", (name,))
            location = _row(
                db.execute("SELECT * FROM locations WHERE id=?", (cursor.lastrowid,)).fetchone()
            )
            store.audit(
                db,
                "CREATE",
                "location",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after=location,
                ip=request.remote_addr,
            )
        return jsonify(location), 201

    @app.route("/Cemes/api/slots", methods=["GET", "POST"])
    @api_guard
    @require_auth
    def cemes_slots():
        if request.method == "GET":
            with store.connection() as db:
                result = list_slots(
                    db,
                    scoped_unit_id(g.cmvr_user),
                    request.args.get("procedure_id"),
                    request.args.get("date_from"),
                    request.args.get("date_to"),
                    request.args.get("available") == "1",
                )
            return jsonify(result)
        if g.cmvr_user["role"] not in ("admin", "regulacao"):
            raise CemesError("Você não possui permissão para esta operação.", 403)
        require_csrf(lambda: None)()
        store.auto_backup()
        with store.connection() as db:
            created = insert_slot(db, body(), g.cmvr_user["id"], request.remote_addr)
        return jsonify(created), 201

    @app.route("/Cemes/api/slots/sequence", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_slots_sequence():
        data = body()
        start = _time(data.get("start_time"))
        end = _time(data.get("end_time"))
        try:
            interval = int(data.get("interval_minutes"))
        except (TypeError, ValueError):
            interval = 0
        if not start or not end or interval not in (10, 15, 20, 30, 45, 60, 90, 120):
            raise CemesError("Sequência de horários inválida.")
        start_min = int(start[:2]) * 60 + int(start[3:])
        end_min = int(end[:2]) * 60 + int(end[3:])
        if end_min < start_min:
            raise CemesError("O horário final deve ser posterior ao inicial.")
        store.auto_backup()
        created = []
        with store.connection() as db:
            for minute in range(start_min, end_min + 1, interval):
                row_data = dict(data)
                row_data["service_time"] = f"{minute // 60:02d}:{minute % 60:02d}"
                row_data["service_time_max"] = row_data["service_time"]
                created.append(
                    insert_slot(db, row_data, g.cmvr_user["id"], request.remote_addr)
                )
        return jsonify(created), 201

    @app.route("/Cemes/api/slots/distribute", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_slots_distribute():
        data = body()
        try:
            procedure_id = int(data.get("procedure_id"))
            doctor_id = int(data.get("doctor_id"))
            total = int(data.get("total_quantity"))
        except (TypeError, ValueError):
            raise CemesError("Informe o procedimento e o total de vagas por data.")
        raw_schedules = data.get("schedules") if isinstance(data.get("schedules"), list) else []
        schedules = [
            {
                "service_date": _date(item.get("service_date")),
                "service_time": _time(item.get("service_time")),
                "service_time_max": _time(item.get("service_time_max") or item.get("service_time")),
            }
            for item in raw_schedules
            if isinstance(item, dict)
        ]
        raw_allocations = data.get("allocations") if isinstance(data.get("allocations"), list) else []
        try:
            allocations = [
                {"unit_id": int(item.get("unit_id")), "quantity": int(item.get("quantity"))}
                for item in raw_allocations
            ]
        except (TypeError, ValueError):
            raise CemesError("A quantidade de cada unidade deve ser um número inteiro igual ou maior que zero.")
        if total < 1 or total > 5000:
            raise CemesError("Informe o procedimento e o total de vagas por data.")
        if (
            not schedules
            or len(schedules) > 60
            or any(not item["service_date"] or not item["service_time"] or not item["service_time_max"] for item in schedules)
        ):
            raise CemesError(
                "Adicione pelo menos uma data com horário inicial e horário máximo válidos. O limite é de 60 agendas."
            )
        if any(item["service_time_max"] < item["service_time"] for item in schedules):
            raise CemesError("O horário máximo não pode ser anterior ao horário inicial.")
        schedule_keys = {
            f"{item['service_date']}|{item['service_time']}|{item['service_time_max']}"
            for item in schedules
        }
        if len(schedule_keys) != len(schedules):
            raise CemesError("Existem datas e horários repetidos na distribuição.")
        store.auto_backup()
        with store.connection() as db:
            active_units = _rows(
                db.execute("SELECT id,name FROM units WHERE active=1 ORDER BY name").fetchall()
            )
            active_ids = {item["id"] for item in active_units}
            allocation_ids = {item["unit_id"] for item in allocations}
            if (
                len(allocations) != len(active_units)
                or len(allocation_ids) != len(active_units)
                or allocation_ids != active_ids
            ):
                raise CemesError(
                    f"A distribuição precisa apresentar todas as {len(active_units)} unidades e setores ativos."
                )
            if any(item["quantity"] < 0 for item in allocations):
                raise CemesError(
                    "A quantidade de cada unidade deve ser um número inteiro igual ou maior que zero."
                )
            allocated = sum(item["quantity"] for item in allocations)
            if allocated < total:
                missing = total - allocated
                raise CemesError(
                    f"A distribuição não foi aceita. {'Falta' if missing == 1 else 'Faltam'} "
                    f"{missing} vaga{'s' if missing != 1 else ''} para completar o total de {total} por data.",
                    400,
                    code="ALLOCATION_MISSING",
                    missing=missing,
                )
            if allocated > total:
                excess = allocated - total
                raise CemesError(
                    f"A distribuição não foi aceita. {'Foi informada' if excess == 1 else 'Foram informadas'} "
                    f"{excess} vaga{'s' if excess != 1 else ''} a mais que o total de {total} por data.",
                    400,
                    code="ALLOCATION_EXCESS",
                    excess=excess,
                )
            procedure = _row(
                db.execute(
                    "SELECT id,name FROM procedures WHERE id=? AND active=1",
                    (procedure_id,),
                ).fetchone()
            )
            if not procedure:
                raise CemesError("Procedimento não encontrado ou inativo.", 404)
            doctor = validate_doctor(db, doctor_id, procedure_id)
            for schedule in schedules:
                for allocation in (item for item in allocations if item["quantity"] > 0):
                    duplicate = _row(
                        db.execute(
                            """
                            SELECT s.id,u.name unit_name FROM slots s
                            JOIN units u ON u.id=s.unit_id
                            WHERE s.unit_id=? AND s.procedure_id=? AND s.service_date=?
                              AND s.service_time<=?
                              AND COALESCE(s.service_time_max,s.service_time)>=?
                              AND s.status='active' LIMIT 1
                            """,
                            (
                                allocation["unit_id"],
                                procedure_id,
                                schedule["service_date"],
                                schedule["service_time_max"],
                                schedule["service_time"],
                            ),
                        ).fetchone()
                    )
                    if duplicate:
                        formatted = "/".join(reversed(schedule["service_date"].split("-")))
                        raise CemesError(
                            f"Já existe uma distribuição de {procedure['name']} para "
                            f"{duplicate['unit_name']} em {formatted} com horário sobreposto. Nada foi gravado.",
                            409,
                        )
            batch_id = f"dist-{int(datetime.now().timestamp() * 1000)}-{secrets.token_hex(4)}"
            created = []
            for schedule in schedules:
                for allocation in (item for item in allocations if item["quantity"] > 0):
                    notes = " — ".join(
                        item
                        for item in (_clean(data.get("notes"), 500), f"Distribuição {batch_id}")
                        if item
                    )
                    created.append(
                        insert_slot(
                            db,
                            {
                                **schedule,
                                "unit_id": allocation["unit_id"],
                                "procedure_id": procedure_id,
                                "quantity": allocation["quantity"],
                                "doctor_id": doctor_id,
                                "location": _clean(data.get("location")) or "CEMES",
                                "notes": notes,
                            },
                            g.cmvr_user["id"],
                            request.remote_addr,
                            do_audit=False,
                        )
                    )
            result = {
                "batch_id": batch_id,
                "procedure_id": procedure_id,
                "procedure_name": procedure["name"],
                "schedules": schedules,
                "dates": [item["service_date"] for item in schedules],
                "dates_count": len(schedules),
                "total_per_date": total,
                "total_distributed": total * len(schedules),
                "allocations": allocations,
                "created_slots": len(created),
                "slots": created,
            }
            store.audit(
                db,
                "DISTRIBUTE",
                "slot_batch",
                batch_id,
                user_id=g.cmvr_user["id"],
                after=result,
                ip=request.remote_addr,
            )
        return jsonify(result), 201

    @app.route("/Cemes/api/slots/<int:slot_id>", methods=["PUT"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_slot_update(slot_id: int):
        data = body()
        store.auto_backup()
        with store.connection() as db:
            before = slot_by_id(db, slot_id)
            if not before:
                raise CemesError("Vaga não encontrada.", 404)
            service_date = _date(data.get("service_date", before["service_date"]))
            service_time = _time(data.get("service_time", before["service_time"]))
            service_time_max = _time(
                data.get("service_time_max", before["service_time_max"] or before["service_time"])
            )
            try:
                quantity = int(data.get("quantity", before["quantity"]))
                unit_id = int(data.get("unit_id", before["unit_id"]))
                procedure_id = int(data.get("procedure_id", before["procedure_id"]))
                doctor_id = int(data.get("doctor_id") or 0) or None
            except (TypeError, ValueError):
                raise CemesError("Dados da vaga inválidos.")
            if (
                not service_date
                or not service_time
                or not service_time_max
                or service_time_max < service_time
                or quantity < max(1, before["used"])
            ):
                raise CemesError(
                    f"A quantidade não pode ser menor que as {before['used']} utilizações já confirmadas."
                )
            provider = _clean(data.get("provider")) or before["provider"]
            if doctor_id:
                provider = validate_doctor(db, doctor_id, procedure_id)["name"]
            db.execute(
                """
                UPDATE slots SET unit_id=?,procedure_id=?,service_date=?,service_time=?,
                service_time_max=?,quantity=?,doctor_id=?,provider=?,location=?,notes=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?
                """,
                (
                    unit_id,
                    procedure_id,
                    service_date,
                    service_time,
                    service_time_max,
                    quantity,
                    doctor_id,
                    provider,
                    _clean(data.get("location")) or before["location"],
                    _clean(data.get("notes"), 500),
                    slot_id,
                ),
            )
            after = slot_by_id(db, slot_id)
            store.audit(
                db,
                "UPDATE",
                "slot",
                slot_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/slots/<int:slot_id>/cancel", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_slot_cancel(slot_id: int):
        reason = _clean(body().get("reason"), 300)
        if not reason:
            raise CemesError("Informe a justificativa do cancelamento.")
        store.auto_backup()
        with store.connection() as db:
            before = slot_by_id(db, slot_id)
            if not before:
                raise CemesError("Vaga não encontrada.", 404)
            if before["used"] > 0:
                raise CemesError("Não é possível cancelar uma vaga com utilização confirmada.", 409)
            note = f"Cancelamento: {reason}"
            db.execute(
                """
                UPDATE slots SET status='cancelled',
                notes=CASE WHEN notes IS NULL OR notes='' THEN ? ELSE notes||char(10)||? END,
                updated_at=CURRENT_TIMESTAMP WHERE id=?
                """,
                (note, note, slot_id),
            )
            after = _row(db.execute("SELECT * FROM slots WHERE id=?", (slot_id,)).fetchone())
            store.audit(
                db,
                "CANCEL",
                "slot",
                slot_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    def can_use_slot(user: dict[str, Any], slot: dict[str, Any]) -> bool:
        scope = user["permissions"]["manual_booking_scope"]
        return (
            scope == "all"
            or (scope == "own" and int(slot["unit_id"]) == int(user["unit_id"]))
            or (
                scope == "own_and_secretaria"
                and (
                    int(slot["unit_id"]) == int(user["unit_id"])
                    or "secretaria" in _norm(slot["unit_name"])
                )
            )
        )

    @app.route("/Cemes/api/slots/<int:slot_id>/use", methods=["POST"])
    @api_guard
    @require_auth
    @require_csrf
    def cemes_slot_use(slot_id: int):
        data = body()
        service_time = _time(data.get("service_time"))
        if not service_time:
            raise CemesError("Informe o horário confirmado no agendamento.")
        store.auto_backup()
        with store.connection() as db:
            slot = slot_by_id(db, slot_id)
            if not slot:
                raise CemesError("Vaga não encontrada.", 404)
            if not can_use_slot(g.cmvr_user, slot):
                raise CemesError(
                    "Este perfil não possui permissão para marcar a utilização desta vaga.", 403
                )
            if slot["status"] != "active":
                raise CemesError("Esta vaga não está ativa.", 409)
            maximum = slot["service_time_max"] or slot["service_time"]
            if service_time < slot["service_time"] or service_time > maximum:
                interval = (
                    slot["service_time"]
                    if maximum == slot["service_time"]
                    else f"{slot['service_time']} até {maximum}"
                )
                raise CemesError(f"O horário deve estar dentro da agenda: {interval}.")
            if slot["used"] >= slot["quantity"]:
                raise CemesError(
                    "Esta vaga já está esgotada e não pode receber outra utilização.", 409
                )
            cursor = db.execute(
                """
                INSERT INTO bookings(
                    slot_id,unit_id,procedure_id,service_date,service_time,
                    operator_id,source,dedupe_key,notes
                ) VALUES(?,?,?,?,?,?,'manual',?,?)
                """,
                (
                    slot_id,
                    slot["unit_id"],
                    slot["procedure_id"],
                    slot["service_date"],
                    service_time,
                    g.cmvr_user["id"],
                    f"manual-slot-{slot_id}-{secrets.token_urlsafe(18)}",
                    _clean(data.get("notes"), 400)
                    or "Agendamento confirmado no portal e registrado manualmente no site.",
                ),
            )
            booking = booking_by_id(db, cursor.lastrowid)
            store.audit(
                db,
                "CONFIRM",
                "booking",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after=booking,
                ip=request.remote_addr,
            )
        return (
            jsonify(
                status="confirmed",
                booking=booking,
                message="Vaga marcada como utilizada e registrada no histórico.",
            ),
            201,
        )

    @app.route("/Cemes/api/slots/<int:slot_id>/transfer", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_slot_transfer(slot_id: int):
        data = body()
        try:
            target_unit = int(data.get("unit_id"))
        except (TypeError, ValueError):
            target_unit = 0
        reason = _clean(data.get("reason"), 300)
        if not target_unit or not reason:
            raise CemesError("Informe a unidade de destino e a justificativa.")
        store.auto_backup()
        with store.connection() as db:
            before = slot_by_id(db, slot_id)
            if not before:
                raise CemesError("Vaga não encontrada.", 404)
            if before["used"] > 0:
                raise CemesError("Não é possível transferir uma vaga que já possui utilização.", 409)
            if not db.execute(
                "SELECT 1 FROM units WHERE id=? AND active=1", (target_unit,)
            ).fetchone():
                raise CemesError("A unidade de destino não existe ou está inativa.")
            note = f"Transferência: {reason}"
            db.execute(
                """
                UPDATE slots SET unit_id=?,
                notes=CASE WHEN notes IS NULL OR notes='' THEN ? ELSE notes||char(10)||? END,
                updated_at=CURRENT_TIMESTAMP WHERE id=?
                """,
                (target_unit, note, note, slot_id),
            )
            after = slot_by_id(db, slot_id)
            store.audit(
                db,
                "TRANSFER",
                "slot",
                slot_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after={**after, "reason": reason},
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/slots/import-xlsx", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_slots_import():
        try:
            workbook = load_workbook(io.BytesIO(request.get_data()), data_only=True)
            sheet = workbook.worksheets[0]
        except Exception:
            raise CemesError("Planilha inválida ou sem conteúdo.")
        headers = {
            str(cell.value or "").strip().lower(): index
            for index, cell in enumerate(sheet[1], start=1)
        }

        def value(row, names):
            key = next((name for name in names if name in headers), None)
            return row[headers[key] - 1].value if key else None

        prepared = []
        errors = []
        with store.connection() as db:
            for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                unit_name = _clean(value(row, ["unidade"]), 160)
                procedure_name = _clean(value(row, ["procedimento"]), 180)
                if not unit_name and not procedure_name:
                    continue
                unit = db.execute(
                    "SELECT id FROM units WHERE name=? COLLATE NOCASE OR short_name=? COLLATE NOCASE",
                    (unit_name, unit_name),
                ).fetchone()
                procedure = db.execute(
                    "SELECT id FROM procedures WHERE name=? COLLATE NOCASE",
                    (procedure_name,),
                ).fetchone()
                if not unit or not procedure:
                    errors.append(
                        {
                            "row": index,
                            "error": "Unidade não encontrada."
                            if not unit
                            else "Procedimento não encontrado.",
                        }
                    )
                    continue
                prepared.append(
                    {
                        "unit_id": unit["id"],
                        "procedure_id": procedure["id"],
                        "service_date": _date(value(row, ["data"])),
                        "service_time": _time(value(row, ["horário", "horario"])),
                        "service_time_max": _time(
                            value(row, ["horário máximo", "horario maximo"])
                            or value(row, ["horário", "horario"])
                        ),
                        "quantity": value(row, ["quantidade"]) or 1,
                        "provider": value(row, ["prestador", "médico", "medico"]),
                        "location": value(row, ["local"]),
                        "notes": value(row, ["observação", "observacao"]),
                    }
                )
        if errors:
            return (
                jsonify(
                    error="Existem linhas que precisam ser corrigidas.",
                    errors=errors,
                    valid_rows=len(prepared),
                ),
                422,
            )
        store.auto_backup()
        created = []
        with store.connection() as db:
            for row_data in prepared:
                created.append(
                    insert_slot(db, row_data, g.cmvr_user["id"], request.remote_addr)
                )
        return jsonify(imported=len(created), slots=created), 201

    @app.route("/Cemes/api/bookings")
    @api_guard
    @require_auth
    def cemes_bookings():
        where = ["1=1"]
        params: list[Any] = []
        unit_id = scoped_unit_id(g.cmvr_user)
        if unit_id:
            where.append("b.unit_id=?")
            params.append(unit_id)
        if request.args.get("date_from"):
            where.append("b.service_date>=?")
            params.append(request.args["date_from"])
        if request.args.get("date_to"):
            where.append("b.service_date<=?")
            params.append(request.args["date_to"])
        if request.args.get("status"):
            where.append("b.status=?")
            params.append(request.args["status"])
        with store.connection() as db:
            result = _rows(
                db.execute(
                    f"""
                    SELECT b.*,u.name unit_name,p.name procedure_name,usr.name operator_name
                    FROM bookings b JOIN units u ON u.id=b.unit_id
                    JOIN procedures p ON p.id=b.procedure_id
                    LEFT JOIN users usr ON usr.id=b.operator_id
                    WHERE {" AND ".join(where)}
                    ORDER BY b.created_at DESC,b.id DESC LIMIT 2000
                    """,
                    params,
                ).fetchall()
            )
        return jsonify(result)

    @app.route("/Cemes/api/bookings/manual", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_booking_manual():
        data = body()
        try:
            unit_id = int(data.get("unit_id"))
        except (TypeError, ValueError):
            raise CemesError("Informe a unidade.")
        store.auto_backup()
        with store.connection() as db:
            result, status = confirm_booking(
                db, data, unit_id, g.cmvr_user["id"], request.remote_addr, "manual"
            )
        return jsonify(result), status

    @app.route("/Cemes/api/bookings/<int:booking_id>/cancel", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_booking_cancel(booking_id: int):
        reason = _clean(body().get("reason"), 300)
        if not reason:
            raise CemesError("Informe o motivo do cancelamento.")
        store.auto_backup()
        with store.connection() as db:
            before = booking_by_id(db, booking_id)
            if not before:
                raise CemesError("Agendamento não encontrado.", 404)
            if before["status"] != "confirmed":
                raise CemesError("Somente um agendamento confirmado pode ser cancelado.", 409)
            db.execute(
                """
                UPDATE bookings SET status='cancelled',cancelled_at=CURRENT_TIMESTAMP,
                cancelled_by=?,cancellation_reason=? WHERE id=?
                """,
                (g.cmvr_user["id"], reason, booking_id),
            )
            after = booking_by_id(db, booking_id)
            store.audit(
                db,
                "CANCEL",
                "booking",
                booking_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/bookings/<int:booking_id>/reschedule", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_booking_reschedule(booking_id: int):
        data = body()
        service_date = _date(data.get("service_date"))
        service_time = _time(data.get("service_time"))
        reason = _clean(data.get("reason"), 300)
        store.auto_backup()
        with store.connection() as db:
            before = booking_by_id(db, booking_id)
            if not before:
                raise CemesError("Agendamento não encontrado.", 404)
            if before["status"] != "confirmed":
                raise CemesError("Somente um agendamento confirmado pode ser remarcado.", 409)
            try:
                procedure_id = int(data.get("procedure_id") or before["procedure_id"])
            except (TypeError, ValueError):
                procedure_id = 0
            if not service_date or not service_time or not reason or not procedure_id:
                raise CemesError("Informe procedimento, nova data, horário e justificativa.")
            slot = _row(
                db.execute(
                    """
                    SELECT s.* FROM slots s
                    WHERE s.unit_id=? AND s.procedure_id=? AND s.service_date=?
                      AND s.service_time<=?
                      AND COALESCE(s.service_time_max,s.service_time)>=?
                      AND s.status='active'
                      AND (SELECT COUNT(*) FROM bookings b
                           WHERE b.slot_id=s.id AND b.status='confirmed')<s.quantity
                    ORDER BY s.id LIMIT 1
                    """,
                    (before["unit_id"], procedure_id, service_date, service_time, service_time),
                ).fetchone()
            )
            if not slot:
                raise CemesError(
                    "A nova data e horário não possuem saldo disponível para esta unidade.", 409
                )
            cursor = db.execute(
                """
                INSERT INTO bookings(
                    slot_id,unit_id,procedure_id,service_date,service_time,
                    operator_id,source,dedupe_key,status,notes
                ) VALUES(?,?,?,?,?,?,'manual',?,'confirmed',?)
                """,
                (
                    slot["id"],
                    before["unit_id"],
                    procedure_id,
                    service_date,
                    service_time,
                    g.cmvr_user["id"],
                    f"reschedule-{booking_id}-{secrets.token_urlsafe(16)}",
                    reason,
                ),
            )
            new_id = cursor.lastrowid
            db.execute(
                """
                UPDATE bookings SET status='rescheduled',cancelled_at=CURRENT_TIMESTAMP,
                cancelled_by=?,cancellation_reason=?,rescheduled_to=? WHERE id=?
                """,
                (g.cmvr_user["id"], reason, new_id, booking_id),
            )
            after = booking_by_id(db, new_id)
            store.audit(
                db,
                "RESCHEDULE",
                "booking",
                booking_id,
                user_id=g.cmvr_user["id"],
                before=before,
                after=after,
                ip=request.remote_addr,
            )
            original = booking_by_id(db, booking_id)
        return jsonify(original=original, booking=after)

    @app.route("/Cemes/api/pending")
    @api_guard
    @require_auth
    def cemes_pending():
        unit_id = scoped_unit_id(g.cmvr_user)
        with store.connection() as db:
            result = _rows(
                db.execute(
                    """
                    SELECT pr.*,u.name unit_name,p.name procedure_name,usr.name operator_name
                    FROM pending_records pr LEFT JOIN units u ON u.id=pr.unit_id
                    LEFT JOIN procedures p ON p.id=pr.procedure_id
                    LEFT JOIN users usr ON usr.id=pr.operator_id
                    WHERE 1=1
                    """
                    + (" AND pr.unit_id=?" if unit_id else "")
                    + " ORDER BY CASE pr.status WHEN 'open' THEN 0 ELSE 1 END,pr.created_at DESC",
                    (unit_id,) if unit_id else (),
                ).fetchall()
            )
        return jsonify(result)

    @app.route("/Cemes/api/pending/<int:pending_id>/resolve", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin", "regulacao")
    @require_csrf
    def cemes_pending_resolve(pending_id: int):
        data = body()
        store.auto_backup()
        with store.connection() as db:
            pending = _row(
                db.execute(
                    "SELECT * FROM pending_records WHERE id=?", (pending_id,)
                ).fetchone()
            )
            if not pending:
                raise CemesError("Pendência não encontrada.", 404)
            if pending["status"] != "open":
                raise CemesError("Pendência já foi tratada.", 409)
            if data.get("action") == "cancel":
                db.execute(
                    """
                    UPDATE pending_records SET status='cancelled',resolution=?,
                    resolved_by=?,resolved_at=CURRENT_TIMESTAMP WHERE id=?
                    """,
                    (
                        _clean(data.get("resolution"), 400) or "Cancelada após conferência.",
                        g.cmvr_user["id"],
                        pending_id,
                    ),
                )
            else:
                service_date = _date(data.get("service_date") or pending["service_date"])
                service_time = _time(data.get("service_time") or pending["service_time"])
                procedure_id = int(data.get("procedure_id") or pending["procedure_id"] or 0)
                slot = _row(
                    db.execute(
                        """
                        SELECT s.* FROM slots s WHERE s.unit_id=? AND s.procedure_id=?
                        AND s.service_date=? AND s.service_time<=?
                        AND COALESCE(s.service_time_max,s.service_time)>=?
                        AND s.status='active'
                        AND (SELECT COUNT(*) FROM bookings b
                        WHERE b.slot_id=s.id AND b.status='confirmed')<s.quantity
                        ORDER BY s.id LIMIT 1
                        """,
                        (
                            pending["unit_id"],
                            procedure_id,
                            service_date,
                            service_time,
                            service_time,
                        ),
                    ).fetchone()
                )
                if not slot:
                    raise CemesError(
                        "Ainda não existe saldo para resolver esta pendência.", 409
                    )
                cursor = db.execute(
                    """
                    INSERT INTO bookings(
                        slot_id,unit_id,procedure_id,service_date,service_time,
                        operator_id,source,dedupe_key,notes
                    ) VALUES(?,?,?,?,?,?,'manual',?,?)
                    """,
                    (
                        slot["id"],
                        pending["unit_id"],
                        procedure_id,
                        service_date,
                        service_time,
                        g.cmvr_user["id"],
                        f"resolved-{pending['dedupe_key']}-{secrets.token_urlsafe(12)}",
                        _clean(data.get("resolution"), 400) or pending["notes"],
                    ),
                )
                db.execute(
                    """
                    UPDATE pending_records SET status='resolved',resolution=?,
                    resolved_by=?,resolved_at=CURRENT_TIMESTAMP WHERE id=?
                    """,
                    (
                        _clean(data.get("resolution"), 400)
                        or f"Vinculada ao agendamento {cursor.lastrowid}.",
                        g.cmvr_user["id"],
                        pending_id,
                    ),
                )
            after = _row(
                db.execute(
                    "SELECT * FROM pending_records WHERE id=?", (pending_id,)
                ).fetchone()
            )
            store.audit(
                db,
                "RESOLVE",
                "pending",
                pending_id,
                user_id=g.cmvr_user["id"],
                before=pending,
                after=after,
                ip=request.remote_addr,
            )
        return jsonify(after)

    @app.route("/Cemes/api/users", methods=["GET", "POST"])
    @api_guard
    @require_auth
    @require_roles("admin")
    def cemes_users():
        if request.method == "GET":
            with store.connection() as db:
                return jsonify(
                    _rows(
                        db.execute(
                            """
                            SELECT usr.id,usr.name,usr.username,usr.role,usr.unit_id,
                            usr.active,usr.must_change_password,u.name unit_name,usr.created_at
                            FROM users usr LEFT JOIN units u ON u.id=usr.unit_id
                            ORDER BY usr.name
                            """
                        ).fetchall()
                    )
                )
        require_csrf(lambda: None)()
        data = body()
        name = _clean(data.get("name"), 140)
        username = _clean(data.get("username"), 80)
        role = _clean(data.get("role"), 20)
        password = str(data.get("password") or "")
        if not name or not username or role not in ("admin", "regulacao", "unidade", "gestor") or len(password) < 8:
            raise CemesError(
                "Preencha nome, usuário, perfil e senha com pelo menos 8 caracteres."
            )
        unit_id = int(data.get("unit_id") or 0) if role == "unidade" else None
        if role == "unidade" and not unit_id:
            raise CemesError("O perfil Unidade precisa estar vinculado a uma unidade.")
        store.auto_backup()
        with store.connection() as db:
            if role == "unidade":
                existing = db.execute(
                    """
                    SELECT username FROM users
                    WHERE unit_id=? AND role='unidade' AND active=1
                    """,
                    (unit_id,),
                ).fetchone()
                if existing:
                    raise CemesError(
                        f"Esta unidade já possui o perfil {existing['username']}. "
                        "Cada unidade pode ter somente um perfil ativo.",
                        409,
                    )
            cursor = db.execute(
                """
                INSERT INTO users(name,username,password_hash,role,unit_id)
                VALUES(?,?,?,?,?)
                """,
                (name, username, generate_password_hash(password), role, unit_id),
            )
            user = _row(
                db.execute(
                    """
                    SELECT usr.id,usr.name,usr.username,usr.role,usr.unit_id,
                    usr.active,usr.must_change_password,u.name unit_name,usr.created_at
                    FROM users usr LEFT JOIN units u ON u.id=usr.unit_id WHERE usr.id=?
                    """,
                    (cursor.lastrowid,),
                ).fetchone()
            )
            store.audit(
                db,
                "CREATE",
                "user",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after=user,
                ip=request.remote_addr,
            )
        return jsonify(user), 201

    @app.route("/Cemes/api/users/change-password", methods=["POST"])
    @api_guard
    @require_auth
    @require_csrf
    def cemes_change_password():
        data = body()
        new_password = str(data.get("new_password") or "")
        if len(new_password) < 8:
            raise CemesError("A nova senha deve ter pelo menos 8 caracteres.")
        store.auto_backup()
        with store.connection() as db:
            current = db.execute(
                "SELECT password_hash FROM users WHERE id=?", (g.cmvr_user["id"],)
            ).fetchone()
            if not check_password_hash(
                current["password_hash"], str(data.get("current_password") or "")
            ):
                raise CemesError("Senha atual incorreta.")
            db.execute(
                "UPDATE users SET password_hash=?,must_change_password=0 WHERE id=?",
                (generate_password_hash(new_password), g.cmvr_user["id"]),
            )
            store.audit(
                db,
                "CHANGE_PASSWORD",
                "user",
                g.cmvr_user["id"],
                user_id=g.cmvr_user["id"],
                ip=request.remote_addr,
            )
        return jsonify(ok=True)

    @app.route("/Cemes/api/audit")
    @api_guard
    @require_auth
    @require_roles("admin")
    def cemes_audit():
        try:
            limit = min(max(int(request.args.get("limit", 200)), 1), 1000)
        except ValueError:
            limit = 200
        with store.connection() as db:
            result = _rows(
                db.execute(
                    """
                    SELECT log.*,usr.name user_name FROM audit_logs log
                    LEFT JOIN users usr ON usr.id=log.user_id
                    ORDER BY log.id DESC LIMIT ?
                    """,
                    (limit,),
                ).fetchall()
            )
        return jsonify(result)

    @app.route("/Cemes/api/admin/devices", methods=["GET", "POST"])
    @api_guard
    @require_auth
    @require_roles("admin")
    def cemes_devices():
        if request.method == "GET":
            with store.connection() as db:
                return jsonify(
                    _rows(
                        db.execute(
                            """
                            SELECT dev.id,dev.name,dev.unit_id,u.name unit_name,
                            dev.operator_id,usr.name operator_name,dev.active,
                            dev.last_seen_at,dev.created_at
                            FROM extension_devices dev JOIN units u ON u.id=dev.unit_id
                            LEFT JOIN users usr ON usr.id=dev.operator_id ORDER BY dev.name
                            """
                        ).fetchall()
                    )
                )
        require_csrf(lambda: None)()
        data = body()
        name = _clean(data.get("name"), 120)
        try:
            unit_id = int(data.get("unit_id"))
        except (TypeError, ValueError):
            unit_id = 0
        if not name or not unit_id:
            raise CemesError("Informe o computador e a unidade.")
        store.auto_backup()
        plain_token = f"cmvr_{secrets.token_urlsafe(24)}"
        token_hash = hashlib.sha256(plain_token.encode()).hexdigest()
        with store.connection() as db:
            unit = db.execute(
                "SELECT id FROM units WHERE id=? AND active=1", (unit_id,)
            ).fetchone()
            if not unit:
                raise CemesError("A extensão precisa estar vinculada a uma unidade ativa.")
            operator = db.execute(
                """
                SELECT id FROM users WHERE unit_id=? AND role='unidade' AND active=1
                ORDER BY id LIMIT 1
                """,
                (unit_id,),
            ).fetchone()
            cursor = db.execute(
                """
                INSERT INTO extension_devices(
                    name,unit_id,operator_id,token_hash,created_by
                ) VALUES(?,?,?,?,?)
                """,
                (
                    name,
                    unit_id,
                    operator["id"] if operator else None,
                    token_hash,
                    g.cmvr_user["id"],
                ),
            )
            store.audit(
                db,
                "CREATE",
                "extension_device",
                cursor.lastrowid,
                user_id=g.cmvr_user["id"],
                after={"name": name, "unit_id": unit_id},
                ip=request.remote_addr,
            )
        return jsonify(id=cursor.lastrowid, token=plain_token), 201

    @app.route("/Cemes/api/admin/backups", methods=["GET", "POST"])
    @api_guard
    @require_auth
    @require_roles("admin")
    def cemes_backups():
        if request.method == "GET":
            return jsonify(store.list_backups())
        require_csrf(lambda: None)()
        backup = store.create_backup("manual")
        with store.connection() as db:
            store.audit(
                db,
                "BACKUP",
                "database",
                backup["name"],
                user_id=g.cmvr_user["id"],
                after=backup,
                ip=request.remote_addr,
            )
        return jsonify(backup), 201

    @app.route("/Cemes/api/admin/backups/<path:name>/download")
    @api_guard
    @require_auth
    @require_roles("admin")
    def cemes_backup_download(name: str):
        path = store.backup_path(name)
        if not path:
            raise CemesError("Backup não encontrado.", 404)
        return send_file(path, as_attachment=True, download_name=path.name)

    @app.route("/Cemes/api/admin/backups/<path:name>/restore", methods=["POST"])
    @api_guard
    @require_auth
    @require_roles("admin")
    @require_csrf
    def cemes_backup_restore(name: str):
        if body().get("confirmation") != "RESTAURAR":
            raise CemesError("Digite RESTAURAR para confirmar.")
        result = store.restore_backup(name)
        with store.connection() as db:
            store.audit(
                db,
                "RESTORE",
                "database",
                name,
                user_id=g.cmvr_user["id"],
                after=result,
                ip=request.remote_addr,
            )
        return jsonify(result)

    def report_rows() -> list[dict[str, Any]]:
        with store.connection() as db:
            return list_slots(
                db,
                scoped_unit_id(g.cmvr_user),
                date_from=request.args.get("date_from"),
                date_to=request.args.get("date_to"),
            )

    @app.route("/Cemes/api/reports/utilization.csv")
    @api_guard
    @require_auth
    def cemes_report_csv():
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            ["Unidade", "Procedimento", "Data", "Horário", "Total", "Utilizadas", "Restantes"]
        )
        for item in report_rows():
            writer.writerow(
                [
                    item["unit_name"],
                    item["procedure_name"],
                    item["service_date"],
                    item["service_time"],
                    item["quantity"],
                    item["used"],
                    item["remaining"],
                ]
            )
        return Response(
            "\ufeff" + output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": 'attachment; filename="relatorio-utilizacao-cmvr.csv"'},
        )

    @app.route("/Cemes/api/reports/utilization.xlsx")
    @api_guard
    @require_auth
    def cemes_report_xlsx():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Utilização"
        sheet.append(
            ["Unidade", "Procedimento", "Data", "Horário", "Total", "Utilizadas", "Restantes"]
        )
        for item in report_rows():
            sheet.append(
                [
                    item["unit_name"],
                    item["procedure_name"],
                    item["service_date"],
                    item["service_time"],
                    item["quantity"],
                    item["used"],
                    item["remaining"],
                ]
            )
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="075985")
        sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
        for column, width in zip("ABCDEFG", (28, 34, 14, 12, 10, 12, 12)):
            sheet.column_dimensions[column].width = width
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return send_file(
            stream,
            as_attachment=True,
            download_name="relatorio-utilizacao-cmvr.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/Cemes/api/reports/utilization.pdf")
    @api_guard
    @require_auth
    def cemes_report_pdf():
        stream = io.BytesIO()
        pdf = canvas.Canvas(stream, pagesize=A4)
        width, height = A4

        def header():
            pdf.setFillColor(HexColor("#075985"))
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(36, height - 42, "Controle Municipal de Vagas da Regulação")
            pdf.setFillColor(HexColor("#334155"))
            pdf.setFont("Helvetica", 9)
            pdf.drawString(
                36,
                height - 58,
                f"Relatório emitido em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            )
            return height - 82

        y = header()
        for item in report_rows():
            if y < 54:
                pdf.showPage()
                y = header()
            pdf.setFillColor(HexColor("#0f172a"))
            pdf.setFont("Helvetica-Bold", 9)
            title = (
                f"{'/'.join(reversed(item['service_date'].split('-')))} "
                f"{item['service_time']} - {item['procedure_name']}"
            )
            pdf.drawString(36, y, title[:96])
            y -= 13
            pdf.setFillColor(HexColor("#475569"))
            pdf.setFont("Helvetica", 8)
            detail = (
                f"{item['unit_name']} | Total: {item['quantity']} | "
                f"Utilizadas: {item['used']} | Restantes: {item['remaining']}"
            )
            pdf.drawString(36, y, detail[:110])
            y -= 19
        pdf.save()
        stream.seek(0)
        return send_file(
            stream,
            as_attachment=True,
            download_name="relatorio-utilizacao-cmvr.pdf",
            mimetype="application/pdf",
        )

    def extension_device() -> tuple[dict[str, Any], sqlite3.Connection]:
        auth = request.headers.get("Authorization", "")
        plain = auth[7:] if auth.startswith("Bearer ") else ""
        token_hash = hashlib.sha256(plain.encode()).hexdigest() if plain else ""
        db = store.connect()
        device = _row(
            db.execute(
                """
                SELECT dev.*,u.name unit_name FROM extension_devices dev
                JOIN units u ON u.id=dev.unit_id
                WHERE dev.token_hash=? AND dev.active=1
                """,
                (token_hash,),
            ).fetchone()
        )
        if not device:
            db.close()
            raise CemesError("Chave da extensão inválida ou desativada.", 401)
        db.execute(
            "UPDATE extension_devices SET last_seen_at=CURRENT_TIMESTAMP WHERE id=?",
            (device["id"],),
        )
        db.commit()
        return device, db

    @app.route("/Cemes/api/extension/context", methods=["GET", "OPTIONS"])
    @api_guard
    def cemes_extension_context():
        if request.method == "OPTIONS":
            return "", 204
        device, db = extension_device()
        try:
            slots = list_slots(
                db,
                device["unit_id"],
                date_from=datetime.now().strftime("%Y-%m-%d"),
                available_only=True,
            )
            procedures = _rows(
                db.execute(
                    """
                    SELECT id,name,sigtap,specialty FROM procedures
                    WHERE active=1 ORDER BY name
                    """
                ).fetchall()
            )
        finally:
            db.close()
        return jsonify(
            device={
                "id": device["id"],
                "name": device["name"],
                "unit_id": device["unit_id"],
                "unit_name": device["unit_name"],
            },
            procedures=procedures,
            slots=slots,
        )

    @app.route("/Cemes/api/extension/bookings", methods=["POST", "OPTIONS"])
    @api_guard
    def cemes_extension_booking():
        if request.method == "OPTIONS":
            return "", 204
        device, db = extension_device()
        try:
            store.auto_backup()
            with db:
                result, status = confirm_booking(
                    db,
                    body(),
                    device["unit_id"],
                    device["operator_id"],
                    request.remote_addr,
                    "extension",
                )
        finally:
            db.close()
        return jsonify(result), status

    @app.route("/cemes")
    @app.route("/cemes/")
    def cemes_lower_redirect():
        return redirect("/Cemes/", code=308)

    @app.route("/Cemes")
    def cemes_slash_redirect():
        return redirect("/Cemes/", code=308)

    @app.route("/Cemes/")
    def cemes_index():
        return send_from_directory(cemes_dir, "index.html")

    @app.route("/Cemes/<path:path>")
    def cemes_static(path: str):
        root = cemes_dir.resolve()
        target = (cemes_dir / path).resolve()
        if target.is_relative_to(root) and target.is_file():
            return send_from_directory(cemes_dir, path)
        return send_from_directory(cemes_dir, "index.html")

    return app
