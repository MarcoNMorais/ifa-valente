from __future__ import annotations

import hashlib
import io
import json
import os
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


TEST_ROOT = tempfile.TemporaryDirectory(prefix="cmvr-integracao-")
TEST_PATH = Path(TEST_ROOT.name)
DATA_PATH = TEST_PATH / "var-data"
CEMES_PATH = DATA_PATH / "cemes"
CIS_PATH = DATA_PATH / "cis"
CIS_PATH.mkdir(parents=True, exist_ok=True)

SENTINEL_PATH = DATA_PATH / "dados-existentes-nao-alterar.txt"
SENTINEL_CONTENT = "DADOS DOS OUTROS SISTEMAS"
SENTINEL_PATH.write_text(SENTINEL_CONTENT, encoding="utf-8")

CIS_DATA_PATH = CIS_PATH / "cis_database.json"
CIS_CONTENT = {
    "versao": 14,
    "atualizadoEm": "2026-07-28T10:00:00",
    "pacientes": [{"id": "preservar-1", "nome": "Registro de teste preservado"}],
    "procedimentos": [],
    "codigos": [],
    "locais": [],
    "users": [
        {
            "id": "admin-existente",
            "user": "admin",
            "pass": "senha-anterior",
            "role": "admin",
            "name": "Administrador existente",
            "active": True,
        }
    ],
    "logs": [],
}
CIS_DATA_PATH.write_text(
    json.dumps(CIS_CONTENT, ensure_ascii=False, indent=2), encoding="utf-8"
)

os.environ["IFA_DATA_DIR"] = str(DATA_PATH)
os.environ["CEMES_DATA_DIR"] = str(CEMES_PATH)
os.environ["CIS_DATA_DIR"] = str(CIS_PATH)
os.environ["SECRET_KEY"] = "segredo-exclusivo-dos-testes"
os.environ["COOKIE_SECURE"] = "0"

from cemes_routes import CemesStore  # noqa: E402
from server import app  # noqa: E402


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class CemesIntegrationTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        app.config.update(TESTING=True)
        cls.admin = app.test_client()
        cls.admin_csrf = cls.login(cls.admin, "admin", "Admin@123")
        cls.units = cls.admin.get("/Cemes/api/units").get_json()
        cls.procedures = cls.admin.get("/Cemes/api/procedures").get_json()
        cls.doctors = cls.admin.get("/Cemes/api/doctors").get_json()
        cls.unit_by_name = {row["name"]: row for row in cls.units}
        cls.procedure_by_name = {row["name"]: row for row in cls.procedures}
        cls.doctor_by_name = {row["name"]: row for row in cls.doctors}
        cls.ifa_database = DATA_PATH / "ifa_valente.db"
        cls.ifa_hash_before_cemes_operations = file_hash(cls.ifa_database)

    @staticmethod
    def login(client, username: str, password: str) -> str:
        response = client.post(
            "/Cemes/api/auth/login",
            json={"username": username, "password": password},
        )
        if response.status_code != 200:
            raise AssertionError(
                f"Falha no login de {username}: {response.status_code} {response.get_data(as_text=True)}"
            )
        session_response = client.get("/Cemes/api/session")
        if session_response.status_code != 200:
            raise AssertionError("Sessão não foi criada.")
        return session_response.get_json()["csrf_token"]

    @staticmethod
    def csrf(token: str) -> dict[str, str]:
        return {"X-CSRF-Token": token}

    @classmethod
    def allocations(cls, values: dict[str, int]) -> list[dict[str, int]]:
        return [
            {"unit_id": unit["id"], "quantity": values.get(unit["name"], 0)}
            for unit in cls.units
        ]

    @classmethod
    def distribute(
        cls,
        procedure_name: str,
        doctor_name: str,
        total: int,
        schedules: list[dict[str, str]],
        values: dict[str, int],
    ):
        return cls.admin.post(
            "/Cemes/api/slots/distribute",
            headers=cls.csrf(cls.admin_csrf),
            json={
                "procedure_id": cls.procedure_by_name[procedure_name]["id"],
                "doctor_id": cls.doctor_by_name[doctor_name]["id"],
                "total_quantity": total,
                "location": "CEMES",
                "schedules": schedules,
                "allocations": cls.allocations(values),
            },
        )

    def test_01_rotas_antigas_e_novo_botao(self):
        root = self.admin.get("/")
        self.assertEqual(root.status_code, 200)
        root_html = root.get_data(as_text=True)
        self.assertIn('href="/Cemes"', root_html)
        self.assertIn("Vagas e marcação", root_html)

        self.assertEqual(self.admin.get("/ifa").status_code, 200)
        self.assertEqual(self.admin.get("/CIS").status_code, 200)
        self.assertEqual(self.admin.get("/EstoqueHospital").status_code, 200)

        cis_status = self.admin.get("/api/cis/status")
        self.assertEqual(cis_status.status_code, 200)
        self.assertTrue(cis_status.get_json()["ok"])

    def test_02_saude_estaticos_e_cancelamento_sem_validacao(self):
        health = self.admin.get("/Cemes/api/health")
        self.assertEqual(health.status_code, 200)
        self.assertEqual(health.get_json()["status"], "ok")

        slash_redirect = self.admin.get("/Cemes", follow_redirects=False)
        self.assertEqual(slash_redirect.status_code, 308)
        self.assertEqual(slash_redirect.headers["Location"], "/Cemes/")
        page = self.admin.get("/Cemes/")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Controle Municipal de Vagas e Marcação", html)
        self.assertIn('id="dialog-close"', html)
        self.assertIn('id="dialog-cancel"', html)
        self.assertIn('id="dialog-close" class="close-btn"', html)
        self.assertIn('<button type="button" id="dialog-cancel"', html)
        self.assertNotIn("Acessos iniciais para teste", html)

        self.assertEqual(self.admin.get("/Cemes/app.js").status_code, 200)
        self.assertEqual(self.admin.get("/Cemes/styles.css").status_code, 200)
        simulator = self.admin.get("/Cemes/simulador.html")
        self.assertEqual(simulator.status_code, 200)
        simulator_html = simulator.get_data(as_text=True)
        self.assertIn("./simulador.css", simulator_html)
        self.assertIn("./simulador.js", simulator_html)
        self.assertIn("Simular agendamento com sucesso", simulator_html)
        self.assertEqual(self.admin.get("/Cemes/simulador.css").status_code, 200)
        self.assertEqual(self.admin.get("/Cemes/simulador.js").status_code, 200)
        lower = self.admin.get("/cemes", follow_redirects=False)
        self.assertEqual(lower.status_code, 308)
        self.assertEqual(lower.headers["Location"], "/Cemes/")

    def test_02_seguranca_login_csrf_e_perfis(self):
        bad_login = app.test_client().post(
            "/Cemes/api/auth/login",
            json={"username": "admin", "password": "senha-incorreta"},
        )
        self.assertEqual(bad_login.status_code, 401)

        no_csrf = self.admin.post(
            "/Cemes/api/locations", json={"name": "Local sem CSRF"}
        )
        self.assertEqual(no_csrf.status_code, 403)

        users = self.admin.get("/Cemes/api/users").get_json()
        unit_profiles = [row for row in users if row["role"] == "unidade" and row["active"]]
        self.assertEqual(len(unit_profiles), 13)
        self.assertEqual(len({row["unit_id"] for row in unit_profiles}), 13)

        gestor = app.test_client()
        gestor_csrf = self.login(gestor, "gestor", "Gestor@123")
        self.assertEqual(gestor.get("/Cemes/api/dashboard").status_code, 200)
        self.assertEqual(
            gestor.post(
                "/Cemes/api/slots",
                headers=self.csrf(gestor_csrf),
                json={},
            ).status_code,
            403,
        )

    def test_03_carga_inicial_medicos_e_procedimentos(self):
        self.assertEqual(len(self.units), 13)
        self.assertEqual(len(self.procedures), 14)
        self.assertEqual(len(self.doctors), 15)
        self.assertEqual(
            self.admin.get("/Cemes/api/locations").get_json()[0]["name"], "CEMES"
        )

        gelson = self.doctor_by_name["GELSON CARNEIRO DA CUNHA"]
        self.assertEqual(len(gelson["procedure_ids"]), 2)
        self.assertIn(
            "Consulta em Ginecologia e Obstetrícia", gelson["procedure_names"]
        )
        self.assertIn("OCI do Colo do Útero", gelson["procedure_names"])

        gastro = self.procedure_by_name["OCI em Gastroenterologia"]
        filtered = self.admin.get(
            f"/Cemes/api/doctors?procedure_id={gastro['id']}"
        ).get_json()
        self.assertEqual(
            {row["name"] for row in filtered},
            {"AUGUSTO CESAR QUINTELA SOUZA", "LUCAS DE OLIVEIRA ALVES"},
        )

    def test_03_edita_inativa_e_reativa_medico_procedimento_e_unidade(self):
        anderson = self.doctor_by_name["ANDERSON TEIVE E ARGOLLO DULTRA"]
        urologia_id = self.procedure_by_name["Consulta em Urologia"]["id"]
        gastro_id = self.procedure_by_name["Consulta em Gastroenterologia"]["id"]

        update_doctor = self.admin.put(
            f"/Cemes/api/doctors/{anderson['id']}",
            headers=self.csrf(self.admin_csrf),
            json={
                "name": anderson["name"],
                "crm": "CRM-BA TESTE",
                "procedure_ids": [urologia_id, gastro_id],
                "active": False,
            },
        )
        self.assertEqual(update_doctor.status_code, 200)
        self.assertFalse(
            any(
                row["id"] == anderson["id"]
                for row in self.admin.get("/Cemes/api/doctors").get_json()
            )
        )
        inactive_doctor = next(
            row
            for row in self.admin.get(
                "/Cemes/api/doctors?include_inactive=1"
            ).get_json()
            if row["id"] == anderson["id"]
        )
        self.assertEqual(inactive_doctor["active"], 0)
        self.assertEqual(set(inactive_doctor["procedure_ids"]), {urologia_id, gastro_id})

        restored_doctor = self.admin.put(
            f"/Cemes/api/doctors/{anderson['id']}",
            headers=self.csrf(self.admin_csrf),
            json={
                "name": anderson["name"],
                "crm": anderson["crm"] or "",
                "procedure_ids": anderson["procedure_ids"],
                "active": True,
            },
        )
        self.assertEqual(restored_doctor.status_code, 200)

        reumatologia = self.procedure_by_name["Consulta em Reumatologia"]
        changed_procedure = self.admin.put(
            f"/Cemes/api/procedures/{reumatologia['id']}",
            headers=self.csrf(self.admin_csrf),
            json={
                "name": "Consulta em Reumatologia — Teste",
                "sigtap": "9999999999",
                "specialty": "Reumatologia",
                "active": False,
            },
        )
        self.assertEqual(changed_procedure.status_code, 200)
        self.assertFalse(
            any(
                row["id"] == reumatologia["id"]
                for row in self.admin.get("/Cemes/api/procedures").get_json()
            )
        )
        CemesStore(app.root_path)
        with sqlite3.connect(CEMES_PATH / "cmvr.db") as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM units").fetchone()[0], 13)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM procedures").fetchone()[0], 14)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM doctors").fetchone()[0], 15)
            self.assertEqual(
                db.execute(
                    "SELECT COUNT(*) FROM procedures WHERE name='Consulta em Reumatologia'"
                ).fetchone()[0],
                0,
            )
        restored_procedure = self.admin.put(
            f"/Cemes/api/procedures/{reumatologia['id']}",
            headers=self.csrf(self.admin_csrf),
            json={
                "name": reumatologia["name"],
                "sigtap": reumatologia["sigtap"],
                "specialty": reumatologia["specialty"],
                "active": True,
            },
        )
        self.assertEqual(restored_procedure.status_code, 200)

        valilandia = self.unit_by_name["USF Valilândia"]
        inactive_unit = self.admin.put(
            f"/Cemes/api/units/{valilandia['id']}",
            headers=self.csrf(self.admin_csrf),
            json={**valilandia, "active": False},
        )
        self.assertEqual(inactive_unit.status_code, 200)
        blocked_profile = app.test_client().post(
            "/Cemes/api/auth/login",
            json={"username": "valilandia", "password": "Unidade@123"},
        )
        self.assertEqual(blocked_profile.status_code, 401)
        restored_unit = self.admin.put(
            f"/Cemes/api/units/{valilandia['id']}",
            headers=self.csrf(self.admin_csrf),
            json={**valilandia, "active": True},
        )
        self.assertEqual(restored_unit.status_code, 200)
        restored_profile = app.test_client().post(
            "/Cemes/api/auth/login",
            json={"username": "valilandia", "password": "Unidade@123"},
        )
        self.assertEqual(restored_profile.status_code, 200)

    def test_04_distribuicao_incompleta_bloqueada(self):
        response = self.distribute(
            "Consulta em Urologia",
            "ANDERSON TEIVE E ARGOLLO DULTRA",
            20,
            [
                {
                    "service_date": "2026-08-19",
                    "service_time": "08:00",
                    "service_time_max": "09:00",
                }
            ],
            {"USF Centro": 19},
        )
        self.assertEqual(response.status_code, 400)
        data = response.get_json()
        self.assertEqual(data["code"], "ALLOCATION_MISSING")
        self.assertEqual(data["missing"], 1)
        self.assertIn("Falta 1 vaga", data["error"])
        self.assertEqual(self.admin.get("/Cemes/api/slots").get_json(), [])

        excess = self.distribute(
            "Consulta em Urologia",
            "ANDERSON TEIVE E ARGOLLO DULTRA",
            20,
            [
                {
                    "service_date": "2026-08-19",
                    "service_time": "08:00",
                    "service_time_max": "09:00",
                }
            ],
            {"USF Centro": 21},
        )
        self.assertEqual(excess.status_code, 400)
        self.assertEqual(excess.get_json()["code"], "ALLOCATION_EXCESS")
        self.assertEqual(excess.get_json()["excess"], 1)
        self.assertIn("1 vaga a mais", excess.get_json()["error"])
        self.assertEqual(self.admin.get("/Cemes/api/slots").get_json(), [])

    def test_05_distribuicao_multiplas_datas_e_horarios(self):
        response = self.distribute(
            "Consulta em Urologia",
            "ANDERSON TEIVE E ARGOLLO DULTRA",
            20,
            [
                {
                    "service_date": "2026-08-19",
                    "service_time": "08:00",
                    "service_time_max": "09:00",
                },
                {
                    "service_date": "2026-08-29",
                    "service_time": "10:00",
                    "service_time_max": "11:30",
                },
            ],
            {"USF Centro": 20},
        )
        self.assertEqual(
            response.status_code, 201, response.get_data(as_text=True)
        )
        data = response.get_json()
        self.assertEqual(data["dates_count"], 2)
        self.assertEqual(data["created_slots"], 2)
        self.assertEqual(data["total_distributed"], 40)

        slots = self.admin.get("/Cemes/api/slots").get_json()
        self.assertEqual(len(slots), 2)
        self.assertEqual(
            {(row["service_date"], row["service_time"], row["service_time_max"]) for row in slots},
            {
                ("2026-08-19", "08:00", "09:00"),
                ("2026-08-29", "10:00", "11:30"),
            },
        )
        dashboard = self.admin.get("/Cemes/api/dashboard").get_json()
        self.assertEqual(dashboard["total"], 40)
        self.assertEqual(dashboard["remaining"], 40)

    def test_06_sobreposicao_e_atomicidade(self):
        response = self.distribute(
            "Consulta em Urologia",
            "ANDERSON TEIVE E ARGOLLO DULTRA",
            20,
            [
                {
                    "service_date": "2026-08-19",
                    "service_time": "08:30",
                    "service_time_max": "09:30",
                }
            ],
            {"USF Centro": 20},
        )
        self.assertEqual(response.status_code, 409)
        self.assertIn("horário sobreposto", response.get_json()["error"])
        self.assertEqual(len(self.admin.get("/Cemes/api/slots").get_json()), 2)

    def test_07_perfil_unidade_visualiza_e_baixa_somente_a_propria_vaga(self):
        center = app.test_client()
        center_csrf = self.login(center, "centro", "Unidade@123")
        session_data = center.get("/Cemes/api/session").get_json()
        self.assertEqual(session_data["user"]["unit_name"], "USF Centro")
        self.assertFalse(session_data["user"]["permissions"]["view_all_units"])

        units = center.get("/Cemes/api/units").get_json()
        self.assertEqual(len(units), 1)
        self.assertEqual(units[0]["name"], "USF Centro")
        slots = center.get("/Cemes/api/slots").get_json()
        self.assertEqual(len(slots), 2)
        self.assertTrue(all(row["unit_name"] == "USF Centro" for row in slots))

        forbidden = center.post(
            "/Cemes/api/slots/distribute",
            headers=self.csrf(center_csrf),
            json={},
        )
        self.assertEqual(forbidden.status_code, 403)

        used = center.post(
            f"/Cemes/api/slots/{slots[0]['id']}/use",
            headers=self.csrf(center_csrf),
            json={"service_time": "08:20"},
        )
        self.assertEqual(used.status_code, 201, used.get_data(as_text=True))
        self.assertEqual(used.get_json()["status"], "confirmed")
        updated = center.get("/Cemes/api/slots").get_json()
        first = next(row for row in updated if row["id"] == slots[0]["id"])
        self.assertEqual(first["used"], 1)
        self.assertEqual(first["remaining"], 19)

    def test_08_perfil_cemes_ve_tudo_e_baixa_cemes_e_secretaria(self):
        response = self.distribute(
            "Ultrassom",
            "ANTONIO EDIL MOTA LOPES",
            3,
            [
                {
                    "service_date": "2026-09-10",
                    "service_time": "14:00",
                    "service_time_max": "15:00",
                }
            ],
            {
                "USF Centro": 1,
                "CEMES": 1,
                "Secretaria Municipal de Saúde": 1,
            },
        )
        self.assertEqual(response.status_code, 201, response.get_data(as_text=True))

        cemes = app.test_client()
        cemes_csrf = self.login(cemes, "cemes", "Unidade@123")
        session_data = cemes.get("/Cemes/api/session").get_json()["user"]
        self.assertTrue(session_data["permissions"]["view_all_units"])
        self.assertEqual(
            session_data["permissions"]["manual_booking_scope"], "own_and_secretaria"
        )

        slots = cemes.get(
            f"/Cemes/api/slots?procedure_id={self.procedure_by_name['Ultrassom']['id']}"
        ).get_json()
        self.assertEqual(len(slots), 3)
        slot_by_unit = {row["unit_name"]: row for row in slots}
        for unit_name in ("CEMES", "Secretaria Municipal de Saúde"):
            result = cemes.post(
                f"/Cemes/api/slots/{slot_by_unit[unit_name]['id']}/use",
                headers=self.csrf(cemes_csrf),
                json={"service_time": "14:30"},
            )
            self.assertEqual(result.status_code, 201, result.get_data(as_text=True))

        forbidden = cemes.post(
            f"/Cemes/api/slots/{slot_by_unit['USF Centro']['id']}/use",
            headers=self.csrf(cemes_csrf),
            json={"service_time": "14:30"},
        )
        self.assertEqual(forbidden.status_code, 403)

    def test_09_extensao_chave_contexto_e_registro(self):
        center_id = self.unit_by_name["USF Centro"]["id"]
        device_response = self.admin.post(
            "/Cemes/api/admin/devices",
            headers=self.csrf(self.admin_csrf),
            json={"name": "Computador USF Centro - Teste", "unit_id": center_id},
        )
        self.assertEqual(
            device_response.status_code, 201, device_response.get_data(as_text=True)
        )
        token = device_response.get_json()["token"]
        bearer = {"Authorization": f"Bearer {token}"}

        context = self.admin.get("/Cemes/api/extension/context", headers=bearer)
        self.assertEqual(context.status_code, 200)
        context_data = context.get_json()
        self.assertEqual(context_data["device"]["unit_name"], "USF Centro")
        self.assertGreaterEqual(len(context_data["slots"]), 1)
        self.assertEqual(len(context_data["procedures"]), 14)

        urologia = self.procedure_by_name["Consulta em Urologia"]
        booking = self.admin.post(
            "/Cemes/api/extension/bookings",
            headers=bearer,
            json={
                "procedure_id": urologia["id"],
                "service_date": "2026-08-29",
                "service_time": "10:45",
                "dedupe_key": "teste-extensao-regulacao-0001",
                "notes": "Solicitação agendada com sucesso.",
            },
        )
        self.assertEqual(booking.status_code, 200, booking.get_data(as_text=True))
        self.assertEqual(booking.get_json()["status"], "confirmed")
        self.__class__.extension_booking_id = booking.get_json()["booking"]["id"]

        duplicate = self.admin.post(
            "/Cemes/api/extension/bookings",
            headers=bearer,
            json={
                "procedure_id": urologia["id"],
                "service_date": "2026-08-29",
                "service_time": "10:45",
                "dedupe_key": "teste-extensao-regulacao-0001",
            },
        )
        self.assertEqual(duplicate.status_code, 200)
        self.assertEqual(duplicate.get_json()["status"], "duplicate")

        invalid = self.admin.get(
            "/Cemes/api/extension/context",
            headers={"Authorization": "Bearer chave-incorreta"},
        )
        self.assertEqual(invalid.status_code, 401)

    def test_10_cancelamento_remarcacao_transferencia_importacao_e_pendencia(self):
        reschedule = self.admin.post(
            f"/Cemes/api/bookings/{self.extension_booking_id}/reschedule",
            headers=self.csrf(self.admin_csrf),
            json={
                "procedure_id": self.procedure_by_name["Consulta em Urologia"]["id"],
                "service_date": "2026-08-19",
                "service_time": "08:40",
                "reason": "Teste de remarcação auditável.",
            },
        )
        self.assertEqual(reschedule.status_code, 200, reschedule.get_data(as_text=True))
        self.assertEqual(reschedule.get_json()["original"]["status"], "rescheduled")
        new_booking_id = reschedule.get_json()["booking"]["id"]

        cancel_booking = self.admin.post(
            f"/Cemes/api/bookings/{new_booking_id}/cancel",
            headers=self.csrf(self.admin_csrf),
            json={"reason": "Teste de devolução de saldo."},
        )
        self.assertEqual(cancel_booking.status_code, 200)
        self.assertEqual(cancel_booking.get_json()["status"], "cancelled")

        ultrassom_slots = self.admin.get(
            f"/Cemes/api/slots?procedure_id={self.procedure_by_name['Ultrassom']['id']}"
        ).get_json()
        center_slot = next(row for row in ultrassom_slots if row["unit_name"] == "USF Centro")
        transfer = self.admin.post(
            f"/Cemes/api/slots/{center_slot['id']}/transfer",
            headers=self.csrf(self.admin_csrf),
            json={
                "unit_id": self.unit_by_name["USF Valilândia"]["id"],
                "reason": "Teste de transferência entre unidades.",
            },
        )
        self.assertEqual(transfer.status_code, 200)
        self.assertEqual(transfer.get_json()["unit_name"], "USF Valilândia")
        transfer_back = self.admin.post(
            f"/Cemes/api/slots/{center_slot['id']}/transfer",
            headers=self.csrf(self.admin_csrf),
            json={
                "unit_id": self.unit_by_name["USF Centro"]["id"],
                "reason": "Restauração após teste.",
            },
        )
        self.assertEqual(transfer_back.status_code, 200)

        workbook = io.BytesIO()
        from openpyxl import Workbook

        source = Workbook()
        sheet = source.active
        sheet.append(
            [
                "Unidade",
                "Procedimento",
                "Data",
                "Horário",
                "Horário máximo",
                "Quantidade",
                "Médico",
                "Local",
                "Observação",
            ]
        )
        sheet.append(
            [
                "CIS",
                "Eletrocardiograma",
                "2026-11-15",
                "07:00",
                "08:00",
                2,
                "MONIA MARIA CARNEIRO GUIMARAES RAMOS",
                "CEMES",
                "Importação de teste",
            ]
        )
        source.save(workbook)
        source.close()
        import_response = self.admin.post(
            "/Cemes/api/slots/import-xlsx",
            headers={
                **self.csrf(self.admin_csrf),
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            data=workbook.getvalue(),
        )
        self.assertEqual(
            import_response.status_code, 201, import_response.get_data(as_text=True)
        )
        self.assertEqual(import_response.get_json()["imported"], 1)

        pending_response = self.admin.post(
            "/Cemes/api/bookings/manual",
            headers=self.csrf(self.admin_csrf),
            json={
                "unit_id": self.unit_by_name["USF Centro"]["id"],
                "procedure_id": self.procedure_by_name["Consulta em Reumatologia"]["id"],
                "service_date": "2026-12-31",
                "service_time": "23:00",
                "dedupe_key": "pendencia-manual-teste-0001",
                "notes": "Teste de pendência.",
            },
        )
        self.assertEqual(pending_response.status_code, 202)
        self.assertEqual(pending_response.get_json()["status"], "pending")
        pending_id = pending_response.get_json()["pending"]["id"]
        resolve = self.admin.post(
            f"/Cemes/api/pending/{pending_id}/resolve",
            headers=self.csrf(self.admin_csrf),
            json={
                "action": "cancel",
                "resolution": "Pendência encerrada durante validação.",
            },
        )
        self.assertEqual(resolve.status_code, 200)
        self.assertEqual(resolve.get_json()["status"], "cancelled")

        cancel_slot_source = self.admin.post(
            "/Cemes/api/slots",
            headers=self.csrf(self.admin_csrf),
            json={
                "unit_id": self.unit_by_name["CIS"]["id"],
                "procedure_id": self.procedure_by_name["Consulta em Reumatologia"]["id"],
                "service_date": "2026-12-20",
                "service_time": "16:00",
                "service_time_max": "17:00",
                "quantity": 1,
                "doctor_id": self.doctor_by_name["ZENON NUNES DA SILVA FILHO"]["id"],
                "location": "CEMES",
            },
        )
        self.assertEqual(cancel_slot_source.status_code, 201)
        cancel_slot = self.admin.post(
            f"/Cemes/api/slots/{cancel_slot_source.get_json()['id']}/cancel",
            headers=self.csrf(self.admin_csrf),
            json={"reason": "Teste de cancelamento de agenda."},
        )
        self.assertEqual(cancel_slot.status_code, 200)
        self.assertEqual(cancel_slot.get_json()["status"], "cancelled")

    def test_11_relatorios_csv_xlsx_pdf(self):
        csv_response = self.admin.get("/Cemes/api/reports/utilization.csv")
        self.assertEqual(csv_response.status_code, 200)
        self.assertIn(
            "text/csv", csv_response.headers.get("Content-Type", "")
        )
        self.assertIn("Consulta em Urologia", csv_response.get_data(as_text=True))

        xlsx_response = self.admin.get("/Cemes/api/reports/utilization.xlsx")
        self.assertEqual(xlsx_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(xlsx_response.data), read_only=True)
        sheet = workbook["Utilização"]
        self.assertEqual(sheet["A1"].value, "Unidade")
        self.assertGreaterEqual(sheet.max_row, 2)
        workbook.close()

        pdf_response = self.admin.get("/Cemes/api/reports/utilization.pdf")
        self.assertEqual(pdf_response.status_code, 200)
        self.assertTrue(pdf_response.data.startswith(b"%PDF"))

    def test_12_backup_auditoria_restauracao_e_persistencia(self):
        backup_response = self.admin.post(
            "/Cemes/api/admin/backups",
            headers=self.csrf(self.admin_csrf),
            json={},
        )
        self.assertEqual(
            backup_response.status_code, 201, backup_response.get_data(as_text=True)
        )
        backup_name = backup_response.get_json()["name"]
        self.assertTrue((CEMES_PATH / "backups" / backup_name).exists())

        listed = self.admin.get("/Cemes/api/admin/backups").get_json()
        self.assertIn(backup_name, {row["name"] for row in listed})
        downloaded = self.admin.get(
            f"/Cemes/api/admin/backups/{backup_name}/download"
        )
        self.assertEqual(downloaded.status_code, 200)
        self.assertGreater(len(downloaded.data), 1024)
        audit_before_restore = self.admin.get("/Cemes/api/audit?limit=1000").get_json()
        self.assertIn("BACKUP", {row["action"] for row in audit_before_restore})

        temporary_location = self.admin.post(
            "/Cemes/api/locations",
            headers=self.csrf(self.admin_csrf),
            json={"name": "LOCAL TEMPORÁRIO PARA TESTE DE RESTAURAÇÃO"},
        )
        self.assertEqual(temporary_location.status_code, 201)
        self.assertTrue(
            any(
                row["name"] == "LOCAL TEMPORÁRIO PARA TESTE DE RESTAURAÇÃO"
                for row in self.admin.get("/Cemes/api/locations").get_json()
            )
        )
        restored = self.admin.post(
            f"/Cemes/api/admin/backups/{backup_name}/restore",
            headers=self.csrf(self.admin_csrf),
            json={"confirmation": "RESTAURAR"},
        )
        self.assertEqual(restored.status_code, 200, restored.get_data(as_text=True))
        self.assertFalse(
            any(
                row["name"] == "LOCAL TEMPORÁRIO PARA TESTE DE RESTAURAÇÃO"
                for row in self.admin.get("/Cemes/api/locations").get_json()
            )
        )

        audit = self.admin.get("/Cemes/api/audit?limit=1000").get_json()
        actions = {row["action"] for row in audit}
        self.assertIn("DISTRIBUTE", actions)
        self.assertIn("CONFIRM", actions)
        self.assertIn("RESTORE", actions)

        database_path = CEMES_PATH / "cmvr.db"
        self.assertTrue(database_path.exists())
        with sqlite3.connect(database_path) as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM procedures").fetchone()[0], 14)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM doctors").fetchone()[0], 15)
            self.assertGreaterEqual(db.execute("SELECT COUNT(*) FROM slots").fetchone()[0], 5)
            self.assertGreaterEqual(db.execute("SELECT COUNT(*) FROM bookings").fetchone()[0], 4)

    def test_13_dados_anteriores_permanecem_intactos(self):
        self.assertEqual(
            SENTINEL_PATH.read_text(encoding="utf-8"), SENTINEL_CONTENT
        )
        self.assertEqual(
            json.loads(CIS_DATA_PATH.read_text(encoding="utf-8")), CIS_CONTENT
        )
        self.assertEqual(
            file_hash(self.ifa_database), self.ifa_hash_before_cemes_operations
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
